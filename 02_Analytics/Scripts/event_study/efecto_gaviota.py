"""
efecto_gaviota.py - Efecto Gaviota inteligente con PAM T8 + series 5 min.

Enfoque:
    PAM Mantto define el evento oficial (fecha + horas_t8).
    La serie temporal define el comportamiento real:
        - inicio del efecto
        - instante de maxima caida
        - recuperacion

Uso standalone:
    python src/efecto_gaviota.py

Uso desde notebook:
    from efecto_gaviota import run_gaviota_analysis
    results = run_gaviota_analysis(...)
"""
from __future__ import annotations

import json
import logging
import re
import unicodedata
import warnings
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import yaml

try:
    import ruptures as rpt

    RUPTURES_AVAILABLE = True
except ImportError:
    rpt = None
    RUPTURES_AVAILABLE = False

warnings.filterwarnings("ignore")
logging.getLogger("matplotlib").setLevel(logging.WARNING)

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_INT = BASE_DIR / "data" / "intermediate"
OUT_FIG = BASE_DIR / "outputs" / "figures"
OUT_XLS = BASE_DIR / "outputs" / "excel"
OUT_RPT = BASE_DIR / "outputs" / "reports"
OUT_GAV = OUT_FIG / "efecto_gaviota"
LOGS_DIR = BASE_DIR / "logs"

for output_dir in (OUT_GAV, OUT_XLS, OUT_RPT, LOGS_DIR):
    output_dir.mkdir(parents=True, exist_ok=True)

TPH_OPERATIONAL = 50
DT_HOURS_5MIN = 5 / 60
ROLLING_POINTS = 12
MIN_PRE_PTS = 50
RECOVERY_SUSTAIN_POINTS = 6
TOP_EVENTOS = 8
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
PAM_SHEET_NAME = "Ejecutivo Mensual"
CANONICAL_DURATION_GROUPS = [2, 4, 8, 12]

# Horarios operacionales oficiales por duración de ventana T8
# Fuente: regla oficial — NO inferir desde datos
T8_WINDOW_TIMES: dict[int, tuple[int, int]] = {
    2:  (14, 16),   # 14:00 → 16:00
    4:  (12, 16),   # 12:00 → 16:00
    8:  ( 8, 16),   # 08:00 → 16:00
    12: ( 8, 20),   # 08:00 → 20:00
}

SEARCH_PRE_HOURS = 48
SEARCH_POST_HOURS = 48
CENTER_PRE_HOURS = 24
CENTER_POST_HOURS = 24
BASELINE_PRE_HOURS = 24
ONSET_DROP_PCT = 0.03
ONSET_Z_THRESHOLD = -1.0

BASE_COLORS = [
    "#4878D0",
    "#EE854A",
    "#6ACC65",
    "#D65F5F",
    "#956CB4",
    "#8C613C",
    "#DC7EC0",
    "#797979",
]

MONTHS_ES = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}

try:
    cfg = yaml.safe_load((BASE_DIR / "config" / "config.yaml").read_text(encoding="utf-8"))
    ACTIVOS = cfg.get("activos", ["SAG1", "SAG2", "PMC", "UNITARIO"])
    COLORES = cfg.get(
        "colores",
        {
            "SAG1": "#1f77b4",
            "SAG2": "#ff7f0e",
            "PMC": "#2ca02c",
            "UNITARIO": "#d62728",
        },
    )
except Exception:
    ACTIVOS = ["SAG1", "SAG2", "PMC", "UNITARIO"]
    COLORES = {
        "SAG1": "#1f77b4",
        "SAG2": "#ff7f0e",
        "PMC": "#2ca02c",
        "UNITARIO": "#d62728",
    }


def normalize_text(value: Any) -> str:
    """Normaliza texto para busquedas robustas."""
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def resolve_existing_path(path_like: str | Path | None, fallbacks: list[Path] | None = None) -> Path | None:
    """Resuelve ruta relativa, absoluta o fallback."""
    candidates: list[Path] = []
    if path_like:
        raw_path = Path(path_like)
        candidates.append(raw_path)
        if not raw_path.is_absolute():
            candidates.append(BASE_DIR / raw_path)
            candidates.append(BASE_DIR / raw_path.name)
    for fallback in fallbacks or []:
        candidates.append(fallback)

    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate)
        if key in seen:
            continue
        seen.add(key)
        if candidate.exists():
            return candidate
    return None


def infer_month_year_from_filename(file_path: str | Path) -> tuple[int | None, int | None]:
    """Infere mes y anio desde el nombre del archivo PAM."""
    file_name = normalize_text(Path(file_path).stem).lower()
    year_match = re.search(r"(20\d{2})", file_name)
    year = int(year_match.group(1)) if year_match else None
    month = next((number for name, number in MONTHS_ES.items() if name in file_name), None)
    return month, year


def load_pam_mantto_files(pam_mantto_dir: str | Path | None = None) -> tuple[list[Path], Path | None, list[str]]:
    """Lista todos los Excel PAM Mantto."""
    notes: list[str] = []
    default_candidates = [
        BASE_DIR / "data" / "raw" / "PAM_Mantto",
        BASE_DIR / "PAM_Mantto",
    ]
    candidate_dirs: list[Path] = []
    if pam_mantto_dir:
        requested = resolve_existing_path(pam_mantto_dir)
        if requested is not None:
            candidate_dirs.append(requested if requested.is_dir() else requested.parent)
        else:
            notes.append(f"Ruta PAM solicitada no existe: {pam_mantto_dir}")

    candidate_dirs.extend(default_candidates)

    seen: set[str] = set()
    normalized_dirs: list[Path] = []
    for candidate in candidate_dirs:
        key = str(candidate.resolve()) if candidate.exists() else str(candidate)
        if key in seen:
            continue
        seen.add(key)
        normalized_dirs.append(candidate)

    for target_dir in normalized_dirs:
        if not target_dir.exists():
            continue
        files = sorted(path for path in target_dir.glob("*") if path.suffix.lower() in EXCEL_SUFFIXES)
        if not files:
            continue
        if pam_mantto_dir and Path(pam_mantto_dir) != target_dir:
            notes.append(f"Se usa PAM Mantto desde: {target_dir}")
        if target_dir == BASE_DIR / "PAM_Mantto":
            notes.append("Se usa fallback legacy 'PAM_Mantto' en raiz del proyecto.")
        return files, target_dir, notes

    return [], None, notes + ["No se encontro directorio PAM Mantto con archivos Excel en rutas oficiales ni fallback."]


def _coerce_header_timestamp(value: Any) -> pd.Timestamp | None:
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return value.normalize()
    if isinstance(value, datetime):
        return pd.Timestamp(value).normalize()
    if isinstance(value, date):
        return pd.Timestamp(datetime.combine(value, datetime.min.time())).normalize()
    if isinstance(value, str):
        raw = value.strip()
        if not raw or not re.search(r"\d", raw):
            return None
        parsed = pd.to_datetime(raw, errors="coerce", dayfirst=True)
        if pd.isna(parsed):
            return None
        return pd.Timestamp(parsed).normalize()
    return None


def _coerce_calendar_day(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, np.integer, float, np.floating)):
        day = int(round(float(value)))
        return day if 1 <= day <= 31 else None
    if isinstance(value, str):
        raw = value.strip()
        if raw.isdigit():
            day = int(raw)
            return day if 1 <= day <= 31 else None
    return None


def detect_t8_row(ws: openpyxl.worksheet.worksheet.Worksheet, max_rows: int = 120) -> tuple[int | None, str | None]:
    """Detecta la fila T8 por texto robusto."""
    target_tokens = ("TENIENTE 8", "VENTANA", "TUNEL PRINCIPAL")
    upper_limit = min(ws.max_row, max_rows)
    for row_idx in range(1, upper_limit + 1):
        row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        row_text = " ".join(filter(None, (normalize_text(value) for value in row_values)))
        if all(token in row_text for token in target_tokens):
            return row_idx, row_text
    return None, None


def extract_calendar_days_from_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    month: int | None,
    year: int | None,
    header_scan_rows: int = 12,
) -> tuple[dict[int, pd.Timestamp], list[str]]:
    """Reconstruye calendario mensual robusto."""
    issues: list[str] = []
    if not month or not year:
        return {}, ["No se pudo inferir mes/anio desde el nombre del archivo."]

    max_col = ws.max_column
    max_row = min(ws.max_row, header_scan_rows)
    calendar_map: dict[int, pd.Timestamp] = {}

    for col_idx in range(1, max_col + 1):
        for row_idx in range(1, max_row + 1):
            timestamp = _coerce_header_timestamp(ws.cell(row=row_idx, column=col_idx).value)
            if timestamp is None:
                continue
            if timestamp.month == month and timestamp.year == year:
                calendar_map[col_idx] = timestamp
                break

    if len(calendar_map) < 10:
        for col_idx in range(1, max_col + 1):
            if col_idx in calendar_map:
                continue
            for row_idx in range(1, max_row + 1):
                day_number = _coerce_calendar_day(ws.cell(row=row_idx, column=col_idx).value)
                if day_number is None:
                    continue
                try:
                    calendar_map[col_idx] = pd.Timestamp(year=year, month=month, day=day_number)
                except ValueError:
                    continue
                break

    if not calendar_map:
        issues.append("No se detectaron columnas de calendario validas.")
        return {}, issues

    if len(calendar_map) < 10:
        issues.append(f"Calendario parcial detectado: solo {len(calendar_map)} columnas con fecha.")

    duplicate_dates = [str(day.date()) for day, count in Counter(calendar_map.values()).items() if count > 1]
    if duplicate_dates:
        issues.append(f"Fechas repetidas en encabezado: {', '.join(duplicate_dates[:5])}")

    invalid_dates = [
        str(value.date())
        for value in calendar_map.values()
        if value.month != month or value.year != year
    ]
    if invalid_dates:
        issues.append(f"Fechas fuera de mes/anio esperado: {', '.join(invalid_dates[:5])}")

    return dict(sorted(calendar_map.items())), issues


def _parse_t8_hours(value: Any) -> float:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return 0.0
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, np.integer, float, np.floating)):
        hours = float(value)
        return hours if hours > 0 else 0.0
    if isinstance(value, str):
        raw = value.strip().replace(",", ".")
        if not raw:
            return 0.0
        try:
            hours = float(raw)
        except ValueError:
            return 0.0
        return hours if hours > 0 else 0.0
    return 0.0


def classify_window_type(hours_t8: float) -> str:
    return f"{int(round(hours_t8))}h"


def duration_group(hours_t8: float) -> int:
    candidates = CANONICAL_DURATION_GROUPS + [int(round(hours_t8))]
    unique_candidates = sorted(set(candidate for candidate in candidates if candidate > 0))
    return min(unique_candidates, key=lambda candidate: abs(candidate - float(hours_t8)))


def get_window_timestamps(fecha: pd.Timestamp, dur_cat: int) -> tuple[pd.Timestamp, pd.Timestamp]:
    """
    Retorna (ini_oficial, fin_oficial) usando los horarios operacionales oficiales.

    La hora de inicio/fin NO se infiere desde los datos; proviene de T8_WINDOW_TIMES.
    Para categorías fuera de la tabla se usa día completo como fallback conservador.
    """
    times = T8_WINDOW_TIMES.get(dur_cat)
    if times is None:
        # fallback: usar el rango más cercano disponible
        closest = min(T8_WINDOW_TIMES.keys(), key=lambda k: abs(k - dur_cat))
        times = T8_WINDOW_TIMES[closest]
    h_ini, h_fin = times
    day = fecha.normalize()
    ini_oficial = day + pd.Timedelta(hours=h_ini)
    fin_oficial  = day + pd.Timedelta(hours=h_fin)
    return ini_oficial, fin_oficial


def extract_t8_events_from_file(file_path: str | Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Extrae serie diaria T8 desde un Excel PAM."""
    path = Path(file_path)
    month, year = infer_month_year_from_filename(path)
    diagnostic: dict[str, Any] = {
        "file": path.name,
        "sheet": None,
        "status": "ok",
        "calendar_issues": [],
        "n_calendar_cols": 0,
        "n_days_extracted": 0,
        "n_positive_events": 0,
        "t8_row": None,
    }

    if month is None or year is None:
        diagnostic["status"] = "filename_problem"
        diagnostic["calendar_issues"] = ["No se pudo inferir mes/anio desde el nombre del archivo."]
        return pd.DataFrame(), diagnostic

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = next((name for name in workbook.sheetnames if normalize_text(name) == normalize_text(PAM_SHEET_NAME)), None)
    diagnostic["sheet"] = sheet_name
    if sheet_name is None:
        workbook.close()
        diagnostic["status"] = "missing_sheet"
        diagnostic["calendar_issues"] = [f"No se encontro hoja '{PAM_SHEET_NAME}'."]
        return pd.DataFrame(), diagnostic

    worksheet = workbook[sheet_name]
    row_idx, _ = detect_t8_row(worksheet)
    diagnostic["t8_row"] = row_idx
    if row_idx is None:
        workbook.close()
        diagnostic["status"] = "missing_t8_row"
        return pd.DataFrame(), diagnostic

    calendar_map, calendar_issues = extract_calendar_days_from_sheet(worksheet, month, year)
    diagnostic["calendar_issues"] = calendar_issues
    diagnostic["n_calendar_cols"] = len(calendar_map)
    if not calendar_map:
        workbook.close()
        diagnostic["status"] = "calendar_problem"
        return pd.DataFrame(), diagnostic

    rows_by_date: dict[pd.Timestamp, dict[str, Any]] = {}
    for col_idx, event_date in calendar_map.items():
        raw_value = worksheet.cell(row=row_idx, column=col_idx).value
        hours_t8 = _parse_t8_hours(raw_value)
        current = rows_by_date.get(event_date)
        record = {
            "fecha": event_date.normalize(),
            "horas_t8": hours_t8,
            "tipo_ventana": classify_window_type(hours_t8) if hours_t8 > 0 else "0h",
            "archivo_origen": path.name,
            "hoja_origen": sheet_name,
        }
        if current is None or hours_t8 > current["horas_t8"]:
            rows_by_date[event_date] = record

    workbook.close()

    if not rows_by_date:
        diagnostic["status"] = "calendar_problem"
        diagnostic["calendar_issues"] = calendar_issues + ["No se pudieron mapear dias del mes a la fila T8."]
        return pd.DataFrame(), diagnostic

    df_file = pd.DataFrame(sorted(rows_by_date.values(), key=lambda item: item["fecha"])).reset_index(drop=True)
    diagnostic["n_days_extracted"] = len(df_file)
    diagnostic["n_positive_events"] = int((df_file["horas_t8"] > 0).sum())
    if calendar_issues:
        diagnostic["status"] = "calendar_warning"
    return df_file, diagnostic


def extract_t8_events_from_folder(pam_mantto_dir: str | Path | None = None) -> tuple[pd.DataFrame, list[dict[str, Any]], dict[str, Any]]:
    """Extrae la serie T8 desde todos los archivos PAM."""
    files, resolved_dir, notes = load_pam_mantto_files(pam_mantto_dir)
    diagnostics: list[dict[str, Any]] = []
    tables: list[pd.DataFrame] = []

    for file_path in files:
        table, diagnostic = extract_t8_events_from_file(file_path)
        diagnostics.append(diagnostic)
        if not table.empty:
            tables.append(table)

    if tables:
        daily_table = pd.concat(tables, ignore_index=True)
    else:
        daily_table = pd.DataFrame(columns=["fecha", "horas_t8", "tipo_ventana", "archivo_origen", "hoja_origen"])

    metadata = {
        "resolved_dir": str(resolved_dir) if resolved_dir else None,
        "n_files_read": len(files),
        "notes": notes,
    }
    return daily_table, diagnostics, metadata


def build_t8_event_table_from_pam(pam_mantto_dir: str | Path | None = None) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    """Construye serie diaria y tabla oficial de eventos T8 > 0."""
    daily_raw, file_diagnostics, metadata = extract_t8_events_from_folder(pam_mantto_dir)

    if daily_raw.empty:
        diagnostics = {
            "resolved_dir": metadata.get("resolved_dir"),
            "n_files_read": metadata.get("n_files_read", 0),
            "notes": metadata.get("notes", []),
            "n_events_detected": 0,
            "date_range": None,
            "duration_distribution": {},
            "files_without_t8_row": [item["file"] for item in file_diagnostics if item["status"] == "missing_t8_row"],
            "files_with_calendar_problems": [
                item["file"] for item in file_diagnostics if item["status"] in {"calendar_problem", "calendar_warning", "filename_problem"}
            ],
            "file_diagnostics": file_diagnostics,
        }
        return daily_raw, daily_raw, diagnostics

    def _join_unique(values: pd.Series) -> str:
        unique_values = sorted({str(value) for value in values if pd.notna(value) and str(value).strip()})
        return " | ".join(unique_values)

    daily_table = (
        daily_raw.groupby("fecha", as_index=False)
        .agg(
            horas_t8=("horas_t8", "max"),
            archivo_origen=("archivo_origen", _join_unique),
            hoja_origen=("hoja_origen", _join_unique),
        )
        .sort_values("fecha")
        .reset_index(drop=True)
    )
    daily_table["fecha"] = pd.to_datetime(daily_table["fecha"]).dt.normalize()
    daily_table["tipo_ventana"] = daily_table["horas_t8"].apply(lambda value: classify_window_type(value) if value > 0 else "0h")
    daily_table["duracion_grupo"] = daily_table["horas_t8"].apply(lambda value: duration_group(value) if value > 0 else 0)

    event_table = daily_table[daily_table["horas_t8"] > 0].copy().reset_index(drop=True)

    diagnostics = {
        "resolved_dir": metadata.get("resolved_dir"),
        "n_files_read": metadata.get("n_files_read", 0),
        "notes": metadata.get("notes", []),
        "n_events_detected": len(event_table),
        "date_range": None if event_table.empty else (str(event_table["fecha"].min().date()), str(event_table["fecha"].max().date())),
        "duration_distribution": {
            str(key): int(value)
            for key, value in event_table["horas_t8"].round(2).value_counts().sort_index().items()
        },
        "files_without_t8_row": [item["file"] for item in file_diagnostics if item["status"] == "missing_t8_row"],
        "files_with_calendar_problems": [
            item["file"] for item in file_diagnostics if item["status"] in {"calendar_problem", "calendar_warning", "filename_problem"}
        ],
        "files_without_sheet": [item["file"] for item in file_diagnostics if item["status"] == "missing_sheet"],
        "file_diagnostics": file_diagnostics,
    }
    return daily_table, event_table, diagnostics


def print_pam_validation_summary(diagnostics: dict[str, Any]) -> None:
    """Imprime validaciones obligatorias."""
    print("  Validacion fuente oficial PAM Mantto")
    print(f"    - Cantidad de archivos PAM Mantto leidos: {diagnostics.get('n_files_read', 0)}")
    print(f"    - Cantidad de eventos T8 detectados: {diagnostics.get('n_events_detected', 0)}")
    print(f"    - Rango de fechas detectado: {diagnostics.get('date_range')}")
    print(f"    - Rango rendimientos disponible: {diagnostics.get('rend_date_range')}")
    print(f"    - Eventos analizables con ventana completa: {diagnostics.get('n_events_analysable')}")
    print(f"    - Eventos oficiales fuera de rango analitico: {diagnostics.get('n_events_outside_rend_range')}")
    print(f"    - Deteccion con ruptures disponible: {diagnostics.get('ruptures_available')}")
    print(f"    - Distribucion de duraciones: {diagnostics.get('duration_distribution', {})}")
    print(f"    - Archivos sin fila T8 encontrada: {diagnostics.get('files_without_t8_row', [])}")
    print(f"    - Archivos con problemas de calendario: {diagnostics.get('files_with_calendar_problems', [])}")
    if diagnostics.get("notes"):
        for note in diagnostics["notes"]:
            print(f"    - Nota: {note}")


def _detect_rend_asset(norm_col: str) -> str | None:
    if "SAG 1" in norm_col or "SAG1" in norm_col:
        return "SAG1"
    if "SAG 2" in norm_col or "SAG2" in norm_col:
        return "SAG2"
    if "CONVENCIONAL" in norm_col or "PMC" in norm_col or "MOL 1 12" in norm_col:
        return "PMC"
    if "MUN" in norm_col or "UNITARIO" in norm_col or "MOL 13" in norm_col:
        return "UNITARIO"
    return None


def _coerce_operando(series: pd.Series) -> pd.Series:
    if pd.api.types.is_bool_dtype(series):
        return series.fillna(False).astype(bool)
    if pd.api.types.is_numeric_dtype(series):
        return series.fillna(0).astype(float) > 0
    normalized = series.fillna("").astype(str).str.strip().str.upper()
    return normalized.isin({"1", "TRUE", "SI", "S", "YES", "Y"})


def normalize_rend_dataframe(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Homologa rendimientos al formato canonico del proyecto."""
    df = df_raw.copy()
    rename_map: dict[str, str] = {}

    for column in df.columns:
        norm_col = normalize_text(column)
        if norm_col.startswith("FECHA"):
            rename_map[column] = "fecha"
            continue
        asset = _detect_rend_asset(norm_col)
        if asset is None:
            continue
        if "OPER" in norm_col:
            rename_map[column] = f"{asset}_operando"
        elif "TON" in norm_col or "TMS" in norm_col:
            rename_map[column] = f"{asset}_ton"
        else:
            rename_map[column] = f"{asset}_tph"

    df = df.rename(columns=rename_map)
    if "fecha" not in df.columns:
        raise ValueError("No se encontro columna de fecha en rendimientos.")

    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    df = df[df["fecha"].notna()].copy()
    df = df.sort_values("fecha").reset_index(drop=True)

    for asset in ACTIVOS:
        tph_col = f"{asset}_tph"
        op_col = f"{asset}_operando"
        ton_col = f"{asset}_ton"

        if tph_col not in df.columns:
            df[tph_col] = np.nan
        df[tph_col] = pd.to_numeric(df[tph_col], errors="coerce")

        if op_col in df.columns:
            df[op_col] = _coerce_operando(df[op_col])
        else:
            df[op_col] = df[tph_col].fillna(0) > TPH_OPERATIONAL

        if ton_col in df.columns:
            df[ton_col] = pd.to_numeric(df[ton_col], errors="coerce").fillna(0)
        else:
            df[ton_col] = df[tph_col].where(df[op_col], 0).fillna(0) * DT_HOURS_5MIN

    ordered = ["fecha"]
    for asset in ACTIVOS:
        ordered.extend([f"{asset}_tph", f"{asset}_operando", f"{asset}_ton"])
    return df[ordered]


def load_rend(rend_path: str | Path | None = None) -> tuple[pd.DataFrame, Path | None, str]:
    """Carga rendimientos 5 min desde parquet o Excel."""
    default_fallbacks = [
        DATA_INT / "rendimientos_clean.parquet",
        BASE_DIR / "data" / "processed" / "fact_rendimiento.parquet",
        BASE_DIR / "rendimientos_coef - copia.xlsx",
    ]
    resolved = resolve_existing_path(rend_path, default_fallbacks)
    if resolved is None:
        raise FileNotFoundError("No se encontro fuente de rendimientos.")

    if resolved.suffix.lower() == ".parquet":
        return normalize_rend_dataframe(pd.read_parquet(resolved)), resolved, "parquet_estandar"

    if resolved.suffix.lower() not in EXCEL_SUFFIXES:
        raise ValueError(f"Formato de rendimientos no soportado: {resolved.suffix}")

    df_excel = pd.read_excel(resolved)
    normalized = normalize_rend_dataframe(df_excel)

    parquet_ref = DATA_INT / "rendimientos_clean.parquet"
    suspicious_assets = []
    for asset in ("PMC", "UNITARIO"):
        q95 = normalized[f"{asset}_tph"].dropna().quantile(0.95) if normalized[f"{asset}_tph"].notna().any() else 0
        if q95 < TPH_OPERATIONAL:
            suspicious_assets.append(asset)

    raw_like = not any(column.endswith("_operando") or column.endswith("_ton") for column in normalized.columns if column != "fecha")
    if suspicious_assets and parquet_ref.exists():
        note = (
            "Excel de rendimientos detectado como fuente cruda no homologada para "
            f"{', '.join(suspicious_assets)}; se usa parquet intermedio estandarizado."
        )
        return normalize_rend_dataframe(pd.read_parquet(parquet_ref)), parquet_ref, note

    if raw_like:
        return normalized, resolved, "Se genero esquema canonico desde Excel crudo usando umbral operacional por defecto."
    return normalized, resolved, "excel_normalizado"


def build_event_records(df_events: pd.DataFrame) -> list[dict[str, Any]]:
    """Convierte tabla oficial de eventos diarios a registros operacionales.

    ini / fin reflejan los horarios operacionales oficiales del T8_WINDOW_TIMES.
    El campo 'fecha' conserva la fecha calendario del PAM (medianoche).
    """
    events: list[dict[str, Any]] = []
    for row in df_events.itertuples(index=False):
        day_start = pd.Timestamp(row.fecha).normalize()
        hours_t8 = float(row.horas_t8)
        dur_cat = int(getattr(row, "duracion_grupo", duration_group(hours_t8)))
        ini_oficial, fin_oficial = get_window_timestamps(day_start, dur_cat)
        events.append(
            {
                "fecha":          day_start,
                "ini":            ini_oficial,   # hora oficial de inicio (ej. 12:00 para 4h)
                "fin":            fin_oficial,   # hora oficial de fin   (ej. 16:00 para 4h)
                "horas_t8":       hours_t8,
                "duracion_cat":   dur_cat,
                "tipo_ventana":   getattr(row, "tipo_ventana", classify_window_type(hours_t8)),
                "archivo_origen": getattr(row, "archivo_origen", ""),
                "hoja_origen":    getattr(row, "hoja_origen", ""),
            }
        )
    return events


def _duration_palette(durations: list[int]) -> dict[int, str]:
    return {duration: BASE_COLORS[index % len(BASE_COLORS)] for index, duration in enumerate(sorted(durations))}


def build_detection_window(
    event: dict[str, Any],
    df_rend: pd.DataFrame,
    search_pre_h: int = SEARCH_PRE_HOURS,
    search_post_h: int = SEARCH_POST_HOURS,
) -> pd.DataFrame:
    """Extrae 48h antes + dia evento + 48h despues."""
    lower = event["ini"] - pd.Timedelta(hours=search_pre_h)
    upper = event["fin"] + pd.Timedelta(hours=search_post_h)
    mask = (df_rend["fecha"] >= lower) & (df_rend["fecha"] <= upper)
    return df_rend.loc[mask].copy()


def _prepare_signal(df_window: pd.DataFrame, asset: str) -> pd.DataFrame:
    """Construye senal con rolling mean/std y relativo al inicio del dia PAM."""
    tph_col = f"{asset}_tph"
    op_col = f"{asset}_operando"
    signal = df_window[["fecha", tph_col, op_col]].copy()
    signal = signal.rename(columns={tph_col: "tph", op_col: "operando"})
    signal["tph"] = pd.to_numeric(signal["tph"], errors="coerce").fillna(0.0)
    signal["operando"] = signal["operando"].fillna(False).astype(bool)
    signal["roll_mean"] = signal["tph"].rolling(ROLLING_POINTS, min_periods=4).mean()
    signal["roll_std"] = signal["tph"].rolling(ROLLING_POINTS, min_periods=4).std()
    return signal


def _detect_change_point(signal: pd.DataFrame, search_start: pd.Timestamp) -> tuple[pd.Timestamp | None, list[pd.Timestamp], str]:
    """Busca quiebres de nivel con ruptures si esta disponible."""
    if signal.empty:
        return None, [], "sin_senal"

    work = signal[["fecha", "roll_mean"]].copy()
    work["roll_mean"] = work["roll_mean"].bfill().ffill().fillna(0.0)
    candidates: list[pd.Timestamp] = []

    if not RUPTURES_AVAILABLE:
        return None, candidates, "rolling_only"

    try:
        values = work["roll_mean"].values.reshape(-1, 1)
        model = rpt.Pelt(model="rbf", min_size=ROLLING_POINTS, jump=1)
        model.fit(values)
        penalty = max(np.nanstd(values) * 2.0, 10.0)
        breakpoints = model.predict(pen=penalty)
        for breakpoint in breakpoints:
            if breakpoint >= len(work):
                continue
            timestamp = pd.Timestamp(work.iloc[breakpoint]["fecha"])
            if timestamp >= search_start:
                candidates.append(timestamp)
    except Exception:
        return None, [], "ruptures_error"

    for candidate in candidates:
        before = work.loc[(work["fecha"] >= candidate - pd.Timedelta(hours=2)) & (work["fecha"] < candidate), "roll_mean"]
        after = work.loc[(work["fecha"] >= candidate) & (work["fecha"] < candidate + pd.Timedelta(hours=2)), "roll_mean"]
        if before.empty or after.empty:
            continue
        if after.mean() < before.mean() * (1 - ONSET_DROP_PCT):
            return candidate, candidates, "ruptures"

    return (candidates[0] if candidates else None), candidates, "ruptures"


def _first_sustained_true(mask: pd.Series, sustain_points: int = RECOVERY_SUSTAIN_POINTS) -> pd.Timestamp | None:
    """Retorna el primer timestamp con condicion sostenida."""
    if mask.empty:
        return None
    sustained = mask.astype(int).rolling(sustain_points, min_periods=sustain_points).sum() >= sustain_points
    hits = sustained[sustained].index
    if len(hits) == 0:
        return None
    return hits[0]


def _detect_effect_onset(
    signal: pd.DataFrame,
    baseline_pre: float,
    baseline_std: float,
    search_start: pd.Timestamp,
) -> tuple[pd.Timestamp | None, pd.Timestamp | None, list[pd.Timestamp], str]:
    """Combina rolling heuristics y ruptures para detectar el inicio del efecto."""
    safe_std = baseline_std if baseline_std and baseline_std > 1e-6 else max(baseline_pre * 0.02, 1.0)
    signal = signal.copy()
    signal["z_baseline"] = (signal["roll_mean"] - baseline_pre) / safe_std
    drop_mask = (
        (signal["fecha"] >= search_start)
        & signal["roll_mean"].notna()
        & (signal["roll_mean"] <= baseline_pre * (1 - ONSET_DROP_PCT))
        & (signal["z_baseline"] <= ONSET_Z_THRESHOLD)
    )
    heuristic_time = _first_sustained_true(pd.Series(drop_mask.values, index=signal["fecha"]), sustain_points=3)
    change_point_time, change_point_candidates, cp_method = _detect_change_point(signal, search_start)

    onset_candidates = [timestamp for timestamp in [heuristic_time, change_point_time] if timestamp is not None]
    if onset_candidates:
        onset_time = min(onset_candidates)
    else:
        onset_time = None

    if onset_time is None:
        method = f"{cp_method}+fallback_valle"
    elif change_point_time is not None and onset_time == change_point_time:
        method = f"{cp_method}+rolling"
    elif heuristic_time is not None and onset_time == heuristic_time:
        method = "rolling"
    else:
        method = cp_method

    return onset_time, heuristic_time, change_point_candidates, method


def _detect_valley(
    signal: pd.DataFrame,
    search_anchor: pd.Timestamp,
    search_end: pd.Timestamp | None = None,
) -> tuple[pd.Timestamp | None, float]:
    """Busca el minimo local de rolling mean dentro del tramo mas plausible del efecto."""
    search_space = signal.loc[signal["fecha"] >= search_anchor, ["fecha", "roll_mean", "tph"]].copy()
    if search_end is not None:
        search_space = search_space.loc[search_space["fecha"] <= search_end].copy()
    if search_space.empty:
        return None, np.nan

    search_space["score"] = search_space["roll_mean"].fillna(search_space["tph"])
    if search_space["score"].isna().all():
        return None, np.nan

    valley_idx = search_space["score"].idxmin()
    valley_time = pd.Timestamp(search_space.loc[valley_idx, "fecha"])
    valley_tph = float(search_space.loc[valley_idx, "score"])
    return valley_time, valley_tph


def _detect_recovery_time(
    signal: pd.DataFrame,
    start_time: pd.Timestamp,
    threshold: float,
    sustain_points: int = RECOVERY_SUSTAIN_POINTS,
) -> pd.Timestamp | None:
    """Detecta cuando la serie vuelve a un porcentaje del baseline."""
    post_signal = signal.loc[signal["fecha"] >= start_time, ["fecha", "roll_mean"]].copy()
    if post_signal.empty:
        return None
    condition = pd.Series((post_signal["roll_mean"] >= threshold).values, index=post_signal["fecha"])
    return _first_sustained_true(condition, sustain_points=sustain_points)


def analyze_event_asset(
    event: dict[str, Any],
    df_rend: pd.DataFrame,
    asset: str,
    search_pre_h: int = SEARCH_PRE_HOURS,
    search_post_h: int = SEARCH_POST_HOURS,
    center_pre_h: int = CENTER_PRE_HOURS,
    center_post_h: int = CENTER_POST_HOURS,
) -> dict[str, Any] | None:
    """Detecta inicio real, minimo y recuperacion para un evento y activo."""
    df_window = build_detection_window(event, df_rend, search_pre_h=search_pre_h, search_post_h=search_post_h)
    if df_window.empty:
        return None

    signal = _prepare_signal(df_window, asset)
    baseline_mask = (
        (signal["fecha"] >= event["ini"] - pd.Timedelta(hours=BASELINE_PRE_HOURS))
        & (signal["fecha"] < event["ini"])
        & signal["operando"]
    )
    baseline_series = signal.loc[baseline_mask, "tph"]
    if len(baseline_series) < MIN_PRE_PTS:
        return None

    baseline_pre = float(baseline_series.mean())
    if baseline_pre < TPH_OPERATIONAL:
        return None
    baseline_std = float(baseline_series.std(ddof=1)) if len(baseline_series) > 1 else max(baseline_pre * 0.02, 1.0)

    onset_time, heuristic_onset, change_point_candidates, detection_method = _detect_effect_onset(
        signal,
        baseline_pre=baseline_pre,
        baseline_std=baseline_std,
        search_start=event["ini"],
    )

    search_anchor = onset_time or event["ini"]
    valley_search_end = min(event["fin"] + pd.Timedelta(hours=24), search_anchor + pd.Timedelta(hours=24))
    valley_time, tph_minimo = _detect_valley(signal, search_anchor=search_anchor, search_end=valley_search_end)
    if valley_time is None:
        valley_time, tph_minimo = _detect_valley(signal, search_anchor=search_anchor, search_end=event["fin"] + pd.Timedelta(hours=48))
    if valley_time is None or pd.isna(tph_minimo):
        return None

    fallback_recovery = signal.loc[signal["fecha"] >= valley_time, "fecha"].max()
    recovery_80 = _detect_recovery_time(signal, valley_time, baseline_pre * 0.80)
    recovery_90 = _detect_recovery_time(signal, valley_time, baseline_pre * 0.90)
    recovery_95 = _detect_recovery_time(signal, valley_time, baseline_pre * 0.95)
    recovery_100 = _detect_recovery_time(signal, valley_time, baseline_pre * 1.00)

    if recovery_100 is None and recovery_95 is not None:
        effect_end = recovery_95
    elif recovery_100 is not None:
        effect_end = recovery_100
    else:
        effect_end = fallback_recovery

    drop_abs = baseline_pre - tph_minimo
    drop_pct = drop_abs / baseline_pre * 100 if baseline_pre > 0 else np.nan

    # t=0 = hora oficial de inicio de la ventana T8 (no el valle)
    ini_oficial = event["ini"]
    center_lower = ini_oficial - pd.Timedelta(hours=center_pre_h)
    center_upper = ini_oficial + pd.Timedelta(hours=center_post_h)
    centered = signal.loc[(signal["fecha"] >= center_lower) & (signal["fecha"] <= center_upper), ["fecha", "tph", "roll_mean"]].copy()
    centered["t_rel_h"] = (centered["fecha"] - ini_oficial).dt.total_seconds() / 3600
    centered["tph_norm_pct"] = centered["tph"] / baseline_pre * 100
    centered["roll_norm_pct"] = centered["roll_mean"] / baseline_pre * 100

    day_mask = (signal["fecha"] >= event["ini"]) & (signal["fecha"] < event["fin"])
    post_day_mask = (signal["fecha"] >= event["fin"]) & (signal["fecha"] < event["fin"] + pd.Timedelta(hours=24))
    tph_event_day = signal.loc[day_mask & signal["operando"], "tph"].mean()
    tph_post_day = signal.loc[post_day_mask & signal["operando"], "tph"].mean()
    delta_post_pre_pct = ((tph_post_day - baseline_pre) / baseline_pre * 100) if pd.notna(tph_post_day) else np.nan

    return {
        "asset": asset,
        "event": event,
        "signal": signal,
        "centered": centered,
        "baseline_pre": baseline_pre,
        "baseline_std": baseline_std,
        "timestamp_inicio_efecto": onset_time,
        "timestamp_inicio_heuristico": heuristic_onset,
        "timestamp_minimo": valley_time,
        "tph_minimo": tph_minimo,
        "timestamp_rec_80": recovery_80,
        "timestamp_rec_90": recovery_90,
        "timestamp_rec_95": recovery_95,
        "timestamp_rec_100": recovery_100,
        "tiempo_hasta_minimo_h": float((valley_time - event["ini"]).total_seconds() / 3600),  # relativo a ini_oficial
        "lag_inicio_efecto_h": float((onset_time - event["ini"]).total_seconds() / 3600) if onset_time is not None else np.nan,
        "tiempo_recuperacion_80_h": float((recovery_80 - valley_time).total_seconds() / 3600) if recovery_80 is not None else np.nan,
        "tiempo_recuperacion_90_h": float((recovery_90 - valley_time).total_seconds() / 3600) if recovery_90 is not None else np.nan,
        "tiempo_recuperacion_95_h": float((recovery_95 - valley_time).total_seconds() / 3600) if recovery_95 is not None else np.nan,
        "tiempo_recuperacion_100_h": float((recovery_100 - valley_time).total_seconds() / 3600) if recovery_100 is not None else np.nan,
        "duracion_efecto_h": float((effect_end - onset_time).total_seconds() / 3600) if onset_time is not None else np.nan,
        "caida_absoluta": drop_abs,
        "caida_pct": drop_pct,
        "delta_post_pre_pct": delta_post_pre_pct,
        "ist8": drop_pct / event["horas_t8"] if event["horas_t8"] > 0 else np.nan,
        "tph_event_day": tph_event_day,
        "tph_post_day": tph_post_day,
        "detection_method": detection_method,
        "change_point_candidates": change_point_candidates,
        "cp_detected": len(change_point_candidates),
    }


def analyze_event(
    event: dict[str, Any],
    df_rend: pd.DataFrame,
    assets: list[str],
    search_pre_h: int = SEARCH_PRE_HOURS,
    search_post_h: int = SEARCH_POST_HOURS,
    center_pre_h: int = CENTER_PRE_HOURS,
    center_post_h: int = CENTER_POST_HOURS,
) -> dict[str, Any]:
    """Analiza un evento oficial T8 para todos los activos."""
    asset_results: dict[str, dict[str, Any]] = {}
    for asset in assets:
        analysis = analyze_event_asset(
            event,
            df_rend,
            asset,
            search_pre_h=search_pre_h,
            search_post_h=search_post_h,
            center_pre_h=center_pre_h,
            center_post_h=center_post_h,
        )
        if analysis is not None:
            asset_results[asset] = analysis
    return {"event": event, "assets": asset_results}


def flatten_event_results(event_results: list[dict[str, Any]]) -> pd.DataFrame:
    """Convierte resultados de deteccion a tabla plana evento-activo."""
    rows: list[dict[str, Any]] = []
    for item in event_results:
        event = item["event"]
        for asset, analysis in item["assets"].items():
            rows.append(
                {
                    "fecha": event["fecha"].date(),
                    "horas_t8": round(event["horas_t8"], 2),
                    "tipo_ventana": event["tipo_ventana"],
                    "duracion_cat": event["duracion_cat"],
                    "archivo_origen": event["archivo_origen"],
                    "activo": asset,
                    "baseline_pre": round(analysis["baseline_pre"], 2),
                    "tph_minimo": round(analysis["tph_minimo"], 2),
                    "caida_absoluta": round(analysis["caida_absoluta"], 2),
                    "caida_pct": round(analysis["caida_pct"], 2),
                    "delta_post_pre_pct": round(analysis["delta_post_pre_pct"], 2) if not pd.isna(analysis["delta_post_pre_pct"]) else np.nan,
                    "timestamp_inicio_efecto": analysis["timestamp_inicio_efecto"],
                    "timestamp_minimo": analysis["timestamp_minimo"],
                    "timestamp_recuperacion_80": analysis["timestamp_rec_80"],
                    "timestamp_recuperacion_90": analysis["timestamp_rec_90"],
                    "timestamp_recuperacion_95": analysis["timestamp_rec_95"],
                    "timestamp_recuperacion_100": analysis["timestamp_rec_100"],
                    "lag_inicio_efecto_h": round(analysis["lag_inicio_efecto_h"], 2) if not pd.isna(analysis["lag_inicio_efecto_h"]) else np.nan,
                    "tiempo_hasta_minimo_h": round(analysis["tiempo_hasta_minimo_h"], 2),
                    "tiempo_recuperacion_80_h": round(analysis["tiempo_recuperacion_80_h"], 2) if not pd.isna(analysis["tiempo_recuperacion_80_h"]) else np.nan,
                    "tiempo_recuperacion_90_h": round(analysis["tiempo_recuperacion_90_h"], 2) if not pd.isna(analysis["tiempo_recuperacion_90_h"]) else np.nan,
                    "tiempo_recuperacion_95_h": round(analysis["tiempo_recuperacion_95_h"], 2) if not pd.isna(analysis["tiempo_recuperacion_95_h"]) else np.nan,
                    "tiempo_recuperacion_100_h": round(analysis["tiempo_recuperacion_100_h"], 2) if not pd.isna(analysis["tiempo_recuperacion_100_h"]) else np.nan,
                    "duracion_efecto_h": round(analysis["duracion_efecto_h"], 2) if not pd.isna(analysis["duracion_efecto_h"]) else np.nan,
                    "ist8": round(analysis["ist8"], 4) if not pd.isna(analysis["ist8"]) else np.nan,
                    "tph_event_day": round(analysis["tph_event_day"], 2) if not pd.isna(analysis["tph_event_day"]) else np.nan,
                    "tph_post_day": round(analysis["tph_post_day"], 2) if not pd.isna(analysis["tph_post_day"]) else np.nan,
                    "detection_method": analysis["detection_method"],
                    "cp_detected": analysis["cp_detected"],
                }
            )
    return pd.DataFrame(rows)


def plot_event_timeline(event_result: dict[str, Any], save: bool = True) -> None:
    """Grafico 1: serie temporal completa con dia PAM, inicio, minimo y recuperacion."""
    event = event_result["event"]
    if not event_result["assets"]:
        return

    fig, axes = plt.subplots(2, 2, figsize=(18, 10), sharex=True)
    fig.suptitle(
        f"Serie temporal completa - evento PAM {event['fecha'].date()} ({event['horas_t8']:.0f}h)",
        fontsize=13,
        fontweight="bold",
    )

    for axis, asset in zip(axes.flat, ACTIVOS):
        analysis = event_result["assets"].get(asset)
        if analysis is None:
            axis.axis("off")
            continue

        signal = analysis["signal"]
        axis.plot(signal["fecha"], signal["tph"], color=COLORES[asset], alpha=0.18, linewidth=0.8, label="TPH")
        axis.plot(signal["fecha"], signal["roll_mean"], color=COLORES[asset], linewidth=2.0, label="Rolling 1h")
        axis.axvspan(event["ini"], event["fin"], color="#D65F5F", alpha=0.10, label="Dia PAM")
        axis.axhline(analysis["baseline_pre"], color="#4C78A8", linestyle=":", linewidth=1.0, label="Baseline pre")

        if analysis["timestamp_inicio_efecto"] is not None:
            axis.axvline(analysis["timestamp_inicio_efecto"], color="#FF8C00", linestyle="--", linewidth=1.2, label="Inicio efecto")
        axis.scatter(analysis["timestamp_minimo"], analysis["tph_minimo"], color="#B22222", s=40, zorder=5, label="Minimo")

        if analysis["timestamp_rec_90"] is not None:
            axis.axvline(analysis["timestamp_rec_90"], color="#2E8B57", linestyle="--", linewidth=1.1, label="Rec 90%")

        axis.set_title(
            f"{asset} | caida {analysis['caida_pct']:.1f}% | rec90 {analysis['tiempo_recuperacion_90_h']:.1f}h"
            if not pd.isna(analysis["tiempo_recuperacion_90_h"])
            else f"{asset} | caida {analysis['caida_pct']:.1f}%"
        )
        axis.grid(True, alpha=0.25)
        axis.set_ylabel("TPH real")
        axis.xaxis.set_major_formatter(mdates.DateFormatter("%d/%m\n%H:%M"))
        axis.legend(fontsize=7, loc="lower right")

    for axis in axes[-1]:
        axis.set_xlabel("Fecha y hora")

    plt.tight_layout(rect=[0, 0, 1, 0.96])
    if save:
        output = OUT_GAV / f"timeline_{event['fecha'].date()}.png"
        fig.savefig(output, dpi=130, bbox_inches="tight")
    plt.close(fig)


def plot_event_gaviota(event_result: dict[str, Any], center_pre_h: int = CENTER_PRE_HOURS, center_post_h: int = CENTER_POST_HOURS, save: bool = True) -> None:
    """Grafico 2: gaviota individual por evento centrada en el minimo."""
    event = event_result["event"]
    if not event_result["assets"]:
        return

    fig, axes = plt.subplots(2, 2, figsize=(18, 10), sharex=True)
    fig.suptitle(
        f"Gaviota individual centrada en la maxima caida - evento {event['fecha'].date()} ({event['horas_t8']:.0f}h)",
        fontsize=13,
        fontweight="bold",
    )

    for axis, asset in zip(axes.flat, ACTIVOS):
        analysis = event_result["assets"].get(asset)
        if analysis is None:
            axis.axis("off")
            continue

        centered = analysis["centered"]
        axis.plot(centered["t_rel_h"], centered["tph"], color=COLORES[asset], alpha=0.18, linewidth=0.8)
        axis.plot(centered["t_rel_h"], centered["roll_mean"], color=COLORES[asset], linewidth=2.0)
        axis.axvline(0, color="#B22222", linestyle="--", linewidth=1.4)
        axis.axhline(analysis["baseline_pre"], color="#4C78A8", linestyle=":", linewidth=1.0)
        axis.scatter(0, analysis["tph_minimo"], color="#B22222", s=40, zorder=5)
        axis.set_xlim(-center_pre_h, center_post_h)
        axis.set_title(f"{asset} | lag {analysis['lag_inicio_efecto_h']:.1f}h | IST8 {analysis['ist8']:.2f}")
        axis.grid(True, alpha=0.25)
        axis.set_ylabel("TPH real")

    for axis in axes[-1]:
        axis.set_xlabel("Horas relativas al minimo")

    plt.tight_layout(rect=[0, 0, 1, 0.96])
    if save:
        output = OUT_GAV / f"gaviota_{event['fecha'].date()}.png"
        fig.savefig(output, dpi=130, bbox_inches="tight")
    plt.close(fig)


def _build_centered_curve_matrices(
    event_results: list[dict[str, Any]],
    center_pre_h: int = CENTER_PRE_HOURS,
    center_post_h: int = CENTER_POST_HOURS,
) -> tuple[np.ndarray, dict[str, dict[int | str, list[np.ndarray]]]]:
    """Construye matrices interpoladas para agregados por activo y duracion."""
    t_axis = np.arange(-center_pre_h, center_post_h + DT_HOURS_5MIN, DT_HOURS_5MIN)
    curves: dict[str, dict[int | str, list[np.ndarray]]] = {asset: {"all": []} for asset in ACTIVOS}

    for event_result in event_results:
        event = event_result["event"]
        duration = int(round(event["horas_t8"]))
        for asset in ACTIVOS:
            analysis = event_result["assets"].get(asset)
            if analysis is None:
                continue
            centered = analysis["centered"][["t_rel_h", "roll_mean"]].dropna()
            if centered.shape[0] < 5:
                continue
            centered = centered.drop_duplicates(subset=["t_rel_h"]).sort_values("t_rel_h")
            interpolated = np.interp(t_axis, centered["t_rel_h"].values, centered["roll_mean"].values, left=np.nan, right=np.nan)
            curves[asset].setdefault(duration, []).append(interpolated)
            curves[asset]["all"].append(interpolated)

    return t_axis, curves


def plot_seagull_by_asset(event_results: list[dict[str, Any]], center_pre_h: int = CENTER_PRE_HOURS, center_post_h: int = CENTER_POST_HOURS) -> dict[str, Any]:
    """
    Figuras 01-04: una figura por activo con todas las duraciones superpuestas.
    t=0 = hora oficial de inicio de la ventana T8.
    Eje Y = TPH real (t/h).
    """
    t_axis, curves = _build_centered_curve_matrices(event_results, center_pre_h=center_pre_h, center_post_h=center_post_h)
    gaviota_data: dict[str, Any] = {}
    dur_palette = _duration_palette(CANONICAL_DURATION_GROUPS)
    DUR_LS = {2: "-", 4: "--", 8: "-.", 12: ":"}

    asset_filenames = {
        "SAG1":     "01_Gaviota_SAG1.png",
        "SAG2":     "02_Gaviota_SAG2.png",
        "PMC":      "03_Gaviota_PMC.png",
        "UNITARIO": "04_Gaviota_UNITARIO.png",
    }

    for idx, asset in enumerate(ACTIVOS, start=1):
        fig, axis = plt.subplots(figsize=(14, 7))
        axis.set_title(
            f"Efecto Gaviota — {asset}\n"
            f"t=0 = inicio oficial ventana T8 | TPH real (t/h) | IQR 25-75 sombreado",
            fontsize=12, fontweight="bold",
        )

        matrix_all = curves.get(asset, {}).get("all", [])
        if not matrix_all:
            axis.text(0.5, 0.5, "Sin datos suficientes", transform=axis.transAxes, ha="center", va="center")
        else:
            # Fondo IQR global
            mat_all = np.vstack(matrix_all)
            p25_all = np.nanpercentile(mat_all, 25, axis=0)
            p75_all = np.nanpercentile(mat_all, 75, axis=0)
            axis.fill_between(t_axis, p25_all, p75_all, color=COLORES[asset], alpha=0.10)

            # Una línea por duración
            for dur in CANONICAL_DURATION_GROUPS:
                dur_curves = curves.get(asset, {}).get(dur, [])
                if not dur_curves:
                    continue
                mat_d = np.vstack(dur_curves)
                med_d = np.nanmedian(mat_d, axis=0)
                axis.plot(
                    t_axis, med_d,
                    color=dur_palette[dur], linewidth=2.2,
                    linestyle=DUR_LS.get(dur, "-"),
                    label=f"T8={dur}h (n={len(dur_curves)})",
                )
                # Marcar la zona de ventana oficial
                dur_times = T8_WINDOW_TIMES.get(dur, (0, dur))
                axis.axvspan(0, dur_times[1] - dur_times[0], color=dur_palette[dur], alpha=0.06)

            # Mediana global
            med_all = np.nanmedian(mat_all, axis=0)
            axis.plot(t_axis, med_all, color=COLORES[asset], linewidth=3.0,
                      linestyle="-", label=f"Mediana global (n={len(matrix_all)})", alpha=0.6)

            gaviota_data[asset] = {
                "t": t_axis, "med": med_all, "p25": p25_all, "p75": p75_all,
                "n": len(matrix_all),
            }

        axis.axvline(0, color="#B22222", linestyle="--", linewidth=1.8, label="Inicio oficial T8")
        axis.set_xlabel("Horas relativas al inicio oficial de la ventana T8", fontsize=10)
        axis.set_ylabel("TPH real (t/h)", fontsize=10)
        axis.set_xlim(-center_pre_h, center_post_h)
        axis.grid(True, alpha=0.25)
        axis.legend(fontsize=9, loc="lower right", framealpha=0.85)

        plt.tight_layout()
        fname = asset_filenames.get(asset, f"0{idx}_Gaviota_{asset}.png")
        fig.savefig(OUT_GAV / fname, dpi=130, bbox_inches="tight")
        plt.close(fig)
        print(f"  {fname}")

    return gaviota_data


def plot_seagull_by_duration(event_results: list[dict[str, Any]], center_pre_h: int = CENTER_PRE_HOURS, center_post_h: int = CENTER_POST_HOURS) -> None:
    """
    Figuras 05-08: una figura por duración T8 con los 4 activos superpuestos.
    t=0 = hora oficial de inicio de la ventana.  Zona sombreada = duración real.
    """
    t_axis = np.arange(-center_pre_h, center_post_h + DT_HOURS_5MIN, DT_HOURS_5MIN)

    dur_filenames = {
        2:  "05_Gaviota_2h.png",
        4:  "06_Gaviota_4h.png",
        8:  "07_Gaviota_8h.png",
        12: "08_Gaviota_12h.png",
    }

    for dur in CANONICAL_DURATION_GROUPS:
        subset = [r for r in event_results if int(round(r["event"]["horas_t8"])) == dur]
        fname = dur_filenames[dur]

        fig, axis = plt.subplots(figsize=(14, 7))
        h_ini, h_fin = T8_WINDOW_TIMES.get(dur, (0, dur))
        dur_h = h_fin - h_ini
        axis.set_title(
            f"Efecto Gaviota — Ventana T8 {dur}h ({h_ini:02d}:00 - {h_fin:02d}:00)\n"
            f"t=0 = {h_ini:02d}:00 (inicio oficial) | todos los activos | TPH real (t/h)",
            fontsize=12, fontweight="bold",
        )

        if not subset:
            axis.text(0.5, 0.5, f"Sin eventos T8={dur}h en el periodo", transform=axis.transAxes,
                      ha="center", va="center", color="gray", fontsize=12)
        else:
            for asset in ACTIVOS:
                asset_curves = []
                for result in subset:
                    analysis = result["assets"].get(asset)
                    if analysis is None:
                        continue
                    centered = analysis["centered"][["t_rel_h", "roll_mean"]].dropna()
                    if centered.shape[0] < 5:
                        continue
                    centered = centered.drop_duplicates(subset=["t_rel_h"]).sort_values("t_rel_h")
                    asset_curves.append(
                        np.interp(t_axis, centered["t_rel_h"].values, centered["roll_mean"].values,
                                  left=np.nan, right=np.nan)
                    )
                if not asset_curves:
                    continue
                mat = np.vstack(asset_curves)
                med = np.nanmedian(mat, axis=0)
                p25 = np.nanpercentile(mat, 25, axis=0)
                p75 = np.nanpercentile(mat, 75, axis=0)
                axis.fill_between(t_axis, p25, p75, color=COLORES[asset], alpha=0.10)
                axis.plot(t_axis, med, color=COLORES[asset], linewidth=2.2,
                          label=f"{asset} (n={len(asset_curves)})")
                # Marcar mínimo
                idx_min = int(np.nanargmin(med))
                axis.annotate(
                    f"{med[idx_min]:.0f} t/h",
                    xy=(t_axis[idx_min], med[idx_min]),
                    xytext=(t_axis[idx_min] + 1.5, med[idx_min] - 5),
                    fontsize=8, color=COLORES[asset],
                    arrowprops=dict(arrowstyle="->", color=COLORES[asset], lw=0.8),
                )

            # Zona sombreada = duración real de la ventana (0 → dur_h)
            axis.axvspan(0, dur_h, color="#D65F5F", alpha=0.08, label=f"Ventana T8 {dur}h")

        axis.axvline(0,     color="#B22222", linestyle="--", linewidth=1.8, label=f"Inicio ({h_ini:02d}:00)")
        axis.axvline(dur_h, color="#228B22", linestyle="--", linewidth=1.4, label=f"Fin ({h_fin:02d}:00)")
        axis.set_xlabel("Horas relativas al inicio oficial de la ventana T8", fontsize=10)
        axis.set_ylabel("TPH real (t/h)", fontsize=10)
        axis.set_xlim(-center_pre_h, center_post_h)
        axis.grid(True, alpha=0.25)
        axis.legend(fontsize=9, loc="lower right", framealpha=0.85)
        plt.tight_layout()
        fig.savefig(OUT_GAV / fname, dpi=130, bbox_inches="tight")
        plt.close(fig)
        n_ev = len(subset)
        print(f"  {fname}  ({n_ev} eventos)")


def plot_asset_overlay(event_results: list[dict[str, Any]], center_pre_h: int = CENTER_PRE_HOURS, center_post_h: int = CENTER_POST_HOURS) -> None:
    """Grafico 5: superposicion de activos en version normalizada."""
    t_axis = np.arange(-center_pre_h, center_post_h + DT_HOURS_5MIN, DT_HOURS_5MIN)
    fig, axis = plt.subplots(figsize=(14, 7))

    for asset in ACTIVOS:
        curves = []
        for result in event_results:
            analysis = result["assets"].get(asset)
            if analysis is None:
                continue
            centered = analysis["centered"][["t_rel_h", "roll_norm_pct"]].dropna()
            if centered.shape[0] < 5:
                continue
            centered = centered.drop_duplicates(subset=["t_rel_h"]).sort_values("t_rel_h")
            curves.append(np.interp(t_axis, centered["t_rel_h"].values, centered["roll_norm_pct"].values, left=np.nan, right=np.nan))
        if not curves:
            continue
        matrix = np.vstack(curves)
        axis.plot(t_axis, np.nanmedian(matrix, axis=0), color=COLORES[asset], linewidth=2.4, label=f"{asset} (n={len(curves)})")

    axis.axvline(0, color="#B22222", linestyle="--", linewidth=1.4)
    axis.axhline(100, color="#4C78A8", linestyle=":", linewidth=1.0)
    axis.set_title("Superposicion de activos - TPH normalizado respecto al baseline pre", fontsize=13, fontweight="bold")
    axis.set_xlabel("Horas relativas al minimo")
    axis.set_ylabel("TPH normalizado (% baseline)")
    axis.grid(True, alpha=0.25)
    axis.legend(fontsize=9, loc="lower right")
    plt.tight_layout()
    output = OUT_GAV / "02_Superposicion_Activos.png"
    fig.savefig(output, dpi=130, bbox_inches="tight")
    plt.close(fig)


def plot_ranking(df_metrics: pd.DataFrame) -> None:
    """Ranking de sensibilidad y recuperacion por activo."""
    if df_metrics.empty:
        return

    ranking = (
        df_metrics.groupby("activo", as_index=False)
        .agg(
            caida_pct=("caida_pct", "mean"),
            ist8=("ist8", "mean"),
            tiempo_recuperacion_90_h=("tiempo_recuperacion_90_h", "mean"),
        )
        .round(2)
    )

    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    specs = [
        ("caida_pct", "Caida % promedio", False),
        ("ist8", "IST8 promedio", False),
        ("tiempo_recuperacion_90_h", "Recuperacion 90% (h)", False),
    ]

    for axis, (column, title, descending) in zip(axes, specs):
        plot_df = ranking.sort_values(column, ascending=not descending).reset_index(drop=True)
        axis.barh(plot_df["activo"], plot_df[column], color="#D65F5F" if column != "tiempo_recuperacion_90_h" else "#4878D0", alpha=0.88)
        axis.set_title(title, fontsize=11, fontweight="bold")
        axis.grid(True, alpha=0.25, axis="x")
        for idx, value in enumerate(plot_df[column]):
            if pd.isna(value):
                continue
            axis.text(value + 0.3, idx, f"{value:.1f}", va="center", ha="left", fontsize=9)

    plt.tight_layout()
    output = OUT_GAV / "09_Ranking_Sensibilidad.png"
    fig.savefig(output, dpi=130, bbox_inches="tight")
    plt.close(fig)
    print(f"  09_Ranking_Sensibilidad.png")


def plot_recuperacion_post_ventana(df_metrics: pd.DataFrame) -> None:
    """
    Figura 10: tiempos de recuperación 80/90/95/100% por activo y tipo de ventana.
    Genera 10_Recuperacion_Post_Ventana.png.
    """
    if df_metrics.empty:
        return

    rec_cols = ["tiempo_recuperacion_80_h", "tiempo_recuperacion_90_h",
                "tiempo_recuperacion_95_h", "tiempo_recuperacion_100_h"]
    rec_labels = ["80%", "90%", "95%", "100%"]
    rec_colors = ["#6ACC65", "#4878D0", "#EE854A", "#D65F5F"]

    fig, axes = plt.subplots(1, 2, figsize=(18, 7))
    fig.suptitle(
        "Recuperacion Post-Ventana T8\nHoras necesarias para recuperar % del baseline previo",
        fontsize=13, fontweight="bold",
    )

    # Panel izquierdo: por activo
    ax = axes[0]
    rec_by_asset = df_metrics.groupby("activo")[rec_cols].mean()
    x = np.arange(len(rec_by_asset))
    w = 0.18
    for k, (col, label, color) in enumerate(zip(rec_cols, rec_labels, rec_colors)):
        vals = rec_by_asset[col].values
        bars = ax.bar(x + k * w, vals, width=w, label=label, color=color, alpha=0.85)
        for bar, v in zip(bars, vals):
            if not np.isnan(v):
                ax.text(bar.get_x() + bar.get_width() / 2, v + 0.3, f"{v:.0f}h",
                        ha="center", va="bottom", fontsize=7)
    ax.set_xticks(x + 1.5 * w)
    ax.set_xticklabels(rec_by_asset.index, fontsize=10)
    ax.set_ylabel("Horas desde el mínimo")
    ax.set_title("Por activo (promedio todos los eventos)", fontsize=10)
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.25, axis="y")

    # Panel derecho: por tipo de ventana (heatmap numérico)
    ax2 = axes[1]
    rec_by_tipo = df_metrics.groupby("tipo_ventana")[rec_cols].mean().round(1)
    rec_by_tipo.columns = rec_labels
    # Grafico de barras agrupadas por tipo
    x2 = np.arange(len(rec_by_tipo))
    for k, (label, color) in enumerate(zip(rec_labels, rec_colors)):
        vals = rec_by_tipo[label].values
        bars = ax2.bar(x2 + k * w, vals, width=w, label=label, color=color, alpha=0.85)
        for bar, v in zip(bars, vals):
            if not np.isnan(v):
                ax2.text(bar.get_x() + bar.get_width() / 2, v + 0.3, f"{v:.0f}h",
                         ha="center", va="bottom", fontsize=7)
    ax2.set_xticks(x2 + 1.5 * w)
    ax2.set_xticklabels(rec_by_tipo.index, fontsize=10)
    ax2.set_ylabel("Horas desde el mínimo")
    ax2.set_title("Por tipo de ventana T8", fontsize=10)
    ax2.legend(fontsize=8)
    ax2.grid(True, alpha=0.25, axis="y")

    plt.tight_layout(rect=[0, 0, 1, 0.94])
    out = OUT_GAV / "10_Recuperacion_Post_Ventana.png"
    fig.savefig(out, dpi=130, bbox_inches="tight")
    plt.close(fig)
    print("  10_Recuperacion_Post_Ventana.png")


def gaviota_data_to_frame(gaviota_data: dict[str, Any]) -> pd.DataFrame:
    """Convierte agregados por activo a tabla exportable."""
    rows: list[dict[str, Any]] = []
    for asset, payload in gaviota_data.items():
        if len(payload.get("med", [])) == 0:
            continue
        min_idx = int(np.nanargmin(payload["med"]))
        rows.append(
            {
                "activo": asset,
                "n_eventos": payload["n"],
                "tph_min_agregado": round(float(np.nanmin(payload["med"])), 2),
                "hora_tph_min_agregado": round(float(payload["t"][min_idx]), 2),
                "tph_median_t0": round(float(np.interp(0, payload["t"], payload["med"])), 2),
                "tph_median_24h_post": round(float(np.interp(24, payload["t"], payload["med"])), 2),
            }
        )
    return pd.DataFrame(rows)


def export_monitoring_outputs(
    daily_t8: pd.DataFrame,
    event_table: pd.DataFrame,
    df_metrics: pd.DataFrame,
    gaviota_data: dict[str, Any],
    diagnostics: dict[str, Any],
) -> None:
    """Exporta tablas oficiales, metricas y diagnosticos."""
    event_xlsx = OUT_XLS / "eventos_t8_desde_pam.xlsx"
    monitoring_xlsx = OUT_XLS / "monitoreo_pre_post_t8.xlsx"

    diag_rows = []
    for item in diagnostics.get("file_diagnostics", []):
        diag_rows.append(
            {
                "archivo": item.get("file"),
                "hoja": item.get("sheet"),
                "status": item.get("status"),
                "fila_t8": item.get("t8_row"),
                "columnas_calendario": item.get("n_calendar_cols"),
                "dias_extraidos": item.get("n_days_extracted"),
                "eventos_positivos": item.get("n_positive_events"),
                "issues": " | ".join(item.get("calendar_issues", [])),
            }
        )
    df_diag = pd.DataFrame(diag_rows)

    with pd.ExcelWriter(event_xlsx, engine="openpyxl") as writer:
        event_table.to_excel(writer, sheet_name="eventos_t8", index=False)
        daily_t8.to_excel(writer, sheet_name="serie_diaria_t8", index=False)
        if not df_diag.empty:
            df_diag.to_excel(writer, sheet_name="diagnostico_archivos", index=False)

    resumen_activo = pd.DataFrame()
    resumen_tipo = pd.DataFrame()
    resumen_evento = pd.DataFrame()
    if not df_metrics.empty:
        resumen_activo = (
            df_metrics.groupby("activo", as_index=False)
            .agg(
                eventos=("fecha", "count"),
                caida_pct_promedio=("caida_pct", "mean"),
                ist8_promedio=("ist8", "mean"),
                lag_inicio_promedio_h=("lag_inicio_efecto_h", "mean"),
                rec90_promedio_h=("tiempo_recuperacion_90_h", "mean"),
            )
            .round(2)
        )
        resumen_tipo = (
            df_metrics.groupby(["activo", "tipo_ventana"], as_index=False)
            .agg(
                eventos=("fecha", "count"),
                caida_pct_promedio=("caida_pct", "mean"),
                tph_min_promedio=("tph_minimo", "mean"),
                rec90_promedio_h=("tiempo_recuperacion_90_h", "mean"),
                ist8_promedio=("ist8", "mean"),
            )
            .round(2)
        )
        resumen_evento = (
            df_metrics.groupby("fecha", as_index=False)
            .agg(
                horas_t8=("horas_t8", "max"),
                caida_pct_promedio=("caida_pct", "mean"),
                rec90_promedio_h=("tiempo_recuperacion_90_h", "mean"),
            )
            .round(2)
        )

    with pd.ExcelWriter(monitoring_xlsx, engine="openpyxl") as writer:
        df_metrics.to_excel(writer, sheet_name="metricas_evento_activo", index=False)
        resumen_activo.to_excel(writer, sheet_name="resumen_activo", index=False)
        resumen_tipo.to_excel(writer, sheet_name="comparativo_tipo_ventana", index=False)
        resumen_evento.to_excel(writer, sheet_name="resumen_evento", index=False)
        gaviota_data_to_frame(gaviota_data).to_excel(writer, sheet_name="efecto_gaviota_agregado", index=False)


def build_executive_summary(event_table: pd.DataFrame, df_metrics: pd.DataFrame, diagnostics: dict[str, Any]) -> str:
    """Genera resumen ejecutivo markdown."""
    lines = [
        "## Resumen Efecto Gaviota Inteligente - PAM T8 + Series Temporales",
        "",
        f"*Generado:* {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "### Validacion fuente oficial",
        f"- Archivos PAM leidos: {diagnostics.get('n_files_read', 0)}",
        f"- Eventos T8 detectados: {diagnostics.get('n_events_detected', 0)}",
        f"- Eventos analizables con ventana completa: {diagnostics.get('n_events_analysable', 0)}",
        f"- Eventos fuera de rango analitico: {diagnostics.get('n_events_outside_rend_range', 0)}",
        f"- Rango PAM: {diagnostics.get('date_range')}",
        f"- Rango rendimientos: {diagnostics.get('rend_date_range')}",
        f"- ruptures disponible: {diagnostics.get('ruptures_available')}",
        f"- Distribucion de duraciones: {diagnostics.get('duration_distribution', {})}",
        "",
    ]

    if event_table.empty or df_metrics.empty:
        lines.extend(
            [
                "### Estado",
                "- No hubo eventos T8 suficientes para construir el analisis inteligente.",
            ]
        )
        return "\n".join(lines)

    resumen_activo = (
        df_metrics.groupby("activo", as_index=False)
        .agg(
            caida_pct=("caida_pct", "mean"),
            rec90=("tiempo_recuperacion_90_h", "mean"),
            ist8=("ist8", "mean"),
            lag_inicio=("lag_inicio_efecto_h", "mean"),
        )
        .round(2)
    )
    worst_drop = resumen_activo.sort_values("caida_pct", ascending=False).iloc[0]
    slow_recovery = resumen_activo.sort_values("rec90", ascending=False).iloc[0]
    worst_event = df_metrics.sort_values("caida_pct", ascending=False).iloc[0]

    lines.extend(
        [
            "### Hallazgos operacionales",
            f"- Activo con mayor caida promedio: **{worst_drop['activo']}** con {worst_drop['caida_pct']:.1f}% de caida.",
            f"- Activo con recuperacion mas lenta: **{slow_recovery['activo']}** con {slow_recovery['rec90']:.1f} h para recuperar 90%.",
            f"- Mayor sensibilidad IST8: **{resumen_activo.sort_values('ist8', ascending=False).iloc[0]['activo']}**.",
            f"- Peor evento observado: **{worst_event['activo']}** el {worst_event['fecha']} con {worst_event['caida_pct']:.1f}% de caida.",
            "",
        ]
    )

    duration_summary = (
        df_metrics.groupby("tipo_ventana", as_index=False)
        .agg(
            eventos=("fecha", "count"),
            caida_pct_promedio=("caida_pct", "mean"),
            rec90_promedio_h=("tiempo_recuperacion_90_h", "mean"),
            lag_inicio_h=("lag_inicio_efecto_h", "mean"),
        )
        .round(2)
        .sort_values("tipo_ventana")
    )
    lines.extend(
        [
            "### Comparativo por tipo de ventana",
            "| Tipo | Eventos | Caida % promedio | Rec 90% (h) | Retardo inicio (h) |",
            "|---|---:|---:|---:|---:|",
        ]
    )
    for row in duration_summary.itertuples(index=False):
        lines.append(
            f"| {row.tipo_ventana} | {row.eventos} | {row.caida_pct_promedio:.1f} | {row.rec90_promedio_h:.1f} | {row.lag_inicio_h:.1f} |"
        )

    lines.extend(
        [
            "",
            "### Interpretacion",
            "- PAM Mantto sigue siendo la fuente oficial del evento T8.",
            "- El centro de la gaviota ya no es medianoche ni una hora arbitraria: se alinea al `timestamp_minimo` observado en la serie.",
            "- El retardo entre el dia PAM y la caida real se estima desde la serie 5 min mediante rolling mean/std y change points.",
            "- Cuando `ruptures` no esta disponible, el pipeline usa un fallback deterministico basado en rolling mean/std y valle local.",
        ]
    )
    return "\n".join(lines)


def append_runtime_log(diagnostics: dict[str, Any], rend_source: Path | None, outputs: list[str]) -> None:
    """Append json line al log tecnico de la corrida."""
    payload = {
        "fecha": datetime.now().isoformat(),
        "script": "src/efecto_gaviota.py",
        "tarea": "Efecto gaviota inteligente con PAM T8 + series temporales",
        "rend_source": str(rend_source) if rend_source else None,
        "pam_source": diagnostics.get("resolved_dir"),
        "n_eventos": diagnostics.get("n_events_detected", 0),
        "n_eventos_analysable": diagnostics.get("n_events_analysable", 0),
        "ruptures_available": diagnostics.get("ruptures_available"),
        "outputs": outputs,
    }
    with open(LOGS_DIR / "skill_audit.log", "a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=False) + "\n")


def run_gaviota_analysis(
    df_rend: pd.DataFrame | None = None,
    rend_path: str | Path | None = None,
    pam_mantto_dir: str | Path | None = None,
    use_pam_t8_events: bool = True,
    activos: list[str] | None = None,
    top_eventos: int = TOP_EVENTOS,
    pre_hours: int = CENTER_PRE_HOURS,
    post_hours: int = CENTER_POST_HOURS,
    search_pre_hours: int = SEARCH_PRE_HOURS,
    search_post_hours: int = SEARCH_POST_HOURS,
    verbose: bool = True,
) -> dict[str, Any]:
    """
    Ejecuta el analisis completo del efecto gaviota inteligente.

    `pre_hours` y `post_hours` definen la ventana relativa para las gaviotas
    centradas en el minimo real.
    """
    activos = activos or ACTIVOS

    if verbose:
        print("=" * 74)
        print("  Efecto Gaviota Inteligente - PAM T8 + Series Temporales")
        print("=" * 74)

    if df_rend is None:
        if verbose:
            print("  Cargando rendimientos 5 min...")
        df_rend, rend_source, rend_note = load_rend(rend_path)
    else:
        df_rend = normalize_rend_dataframe(df_rend)
        rend_source = None
        rend_note = "DataFrame entregado por el usuario"

    if verbose:
        print(f"  Rendimientos: {len(df_rend):,} registros | {df_rend['fecha'].min()} -> {df_rend['fecha'].max()}")
        print(f"  Fuente rendimientos: {rend_note}")

    if not use_pam_t8_events:
        raise ValueError("Este analisis inteligente requiere PAM Mantto como fuente oficial de eventos T8.")

    daily_t8, event_table, diagnostics = build_t8_event_table_from_pam(pam_mantto_dir)
    diagnostics["ruptures_available"] = RUPTURES_AVAILABLE

    rend_min = df_rend["fecha"].min()
    rend_max = df_rend["fecha"].max()
    diagnostics["rend_date_range"] = (str(rend_min), str(rend_max))

    if verbose:
        print_pam_validation_summary(diagnostics)

    if event_table.empty:
        # Fallback: usar ventanas_t8.parquet si existe
        parquet_vent = DATA_INT / "ventanas_t8.parquet"
        if parquet_vent.exists():
            if verbose:
                print("  PAM Mantto sin eventos — usando fallback: ventanas_t8.parquet")
            df_vent = pd.read_parquet(parquet_vent)
            df_vent["fecha"] = pd.to_datetime(df_vent["fecha"]).dt.normalize()
            df_vent = df_vent[df_vent["horas_t8"] > 0].copy()
            df_vent["tipo_ventana"]  = df_vent["horas_t8"].apply(classify_window_type)
            df_vent["duracion_grupo"] = df_vent["horas_t8"].apply(duration_group)
            df_vent["archivo_origen"] = "ventanas_t8.parquet"
            df_vent["hoja_origen"]    = "parquet"
            daily_t8    = df_vent
            event_table = df_vent
            diagnostics["n_events_detected"]    = len(event_table)
            diagnostics["date_range"]           = (str(event_table["fecha"].min().date()),
                                                   str(event_table["fecha"].max().date()))
            diagnostics["duration_distribution"] = {
                str(k): int(v)
                for k, v in event_table["horas_t8"].round(2).value_counts().sort_index().items()
            }
            if verbose:
                print(f"  Fallback: {len(event_table)} eventos desde parquet")
        else:
            raise ValueError("DETENER ANALISIS: no se detectaron eventos T8 desde PAM Mantto y no existe ventanas_t8.parquet como fallback.")

    # Reconstruir events con los horarios oficiales
    events = build_event_records(event_table)
    rend_min = df_rend["fecha"].min()
    rend_max = df_rend["fecha"].max()
    analysable_events = [
        ev for ev in events
        if ev["ini"] - pd.Timedelta(hours=search_pre_hours) >= rend_min
        and ev["fin"] + pd.Timedelta(hours=search_post_hours) <= rend_max
    ]
    diagnostics["n_events_analysable"] = len(analysable_events)
    diagnostics["n_events_outside_rend_range"] = max(len(events) - len(analysable_events), 0)

    if not analysable_events:
        raise ValueError("DETENER ANALISIS: no hay eventos T8 con 48h antes + dia evento + 48h despues dentro del rango de rendimientos.")

    event_results = [
        analyze_event(
            event,
            df_rend,
            activos,
            search_pre_h=search_pre_hours,
            search_post_h=search_post_hours,
            center_pre_h=pre_hours,
            center_post_h=post_hours,
        )
        for event in analysable_events
    ]
    event_results = [item for item in event_results if item["assets"]]

    df_metrics = flatten_event_results(event_results)
    if verbose:
        print(f"  Metricas inteligentes calculadas: {len(df_metrics)} filas")

    if df_metrics.empty:
        raise ValueError("DETENER ANALISIS: no hubo datos suficientes para detectar caidas y recuperaciones.")

    highlighted_dates = (
        df_metrics.groupby("fecha", as_index=False)["caida_pct"]
        .mean()
        .sort_values("caida_pct", ascending=False)
        .head(top_eventos)["fecha"]
        .tolist()
    )
    highlighted_set = {pd.Timestamp(value).date() for value in highlighted_dates}
    highlighted_events = [result for result in event_results if result["event"]["fecha"].date() in highlighted_set]
    if not highlighted_events:
        highlighted_events = event_results[:top_eventos]

    if verbose:
        print(f"  Graficos individuales: {len(highlighted_events)} eventos")
    for event_result in highlighted_events:
        plot_event_timeline(event_result, save=True)
        plot_event_gaviota(event_result, center_pre_h=pre_hours, center_post_h=post_hours, save=True)

    if verbose:
        print("\n  Generando figuras de Efecto Gaviota...")
    gaviota_data = plot_seagull_by_asset(event_results, center_pre_h=pre_hours, center_post_h=post_hours)
    plot_seagull_by_duration(event_results, center_pre_h=pre_hours, center_post_h=post_hours)
    plot_asset_overlay(event_results, center_pre_h=pre_hours, center_post_h=post_hours)
    plot_ranking(df_metrics)
    plot_recuperacion_post_ventana(df_metrics)
    export_monitoring_outputs(daily_t8, event_table, df_metrics, gaviota_data, diagnostics)

    summary = build_executive_summary(event_table, df_metrics, diagnostics)
    summary_path = OUT_RPT / "resumen_efecto_gaviota.md"
    summary_path.write_text(summary, encoding="utf-8")

    outputs = [
        str(OUT_XLS / "eventos_t8_desde_pam.xlsx"),
        str(OUT_XLS / "monitoreo_pre_post_t8.xlsx"),
        str(OUT_GAV / "01_Gaviota_SAG1.png"),
        str(OUT_GAV / "02_Gaviota_SAG2.png"),
        str(OUT_GAV / "03_Gaviota_PMC.png"),
        str(OUT_GAV / "04_Gaviota_UNITARIO.png"),
        str(OUT_GAV / "05_Gaviota_2h.png"),
        str(OUT_GAV / "06_Gaviota_4h.png"),
        str(OUT_GAV / "07_Gaviota_8h.png"),
        str(OUT_GAV / "08_Gaviota_12h.png"),
        str(OUT_GAV / "09_Ranking_Sensibilidad.png"),
        str(OUT_GAV / "10_Recuperacion_Post_Ventana.png"),
        str(summary_path),
    ]
    append_runtime_log(diagnostics, rend_source, outputs)

    if verbose:
        print(f"  Reporte: {summary_path}")
        print("  Completado.")
        print("=" * 74)

    return {
        "daily_t8": daily_t8,
        "eventos": event_results,
        "eventos_oficiales": events,
        "event_table": event_table,
        "df_met": df_metrics,
        "gaviota_data": gaviota_data,
        "diagnostics": diagnostics,
        "summary": summary,
        "rend_source": str(rend_source) if rend_source else None,
    }


if __name__ == "__main__":
    run_gaviota_analysis(
        rend_path=DATA_INT / "rendimientos_clean.parquet",
        pam_mantto_dir=BASE_DIR / "data" / "raw" / "PAM_Mantto",
        use_pam_t8_events=True,
    )
