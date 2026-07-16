"""
loader.py — Carga de datos PAM Producción, PAM Mantto y Rendimientos.
Detecta automáticamente rutas legacy (raíz) y nueva estructura (data/raw/).
"""

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
import pandas as pd
import yaml

log = logging.getLogger(__name__)

MESES_ES = {
    'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
    'julio':7,'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12
}


def cargar_config(ruta_config: Path) -> dict:
    with open(ruta_config, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def _mes_anio(nombre: str):
    nombre = nombre.lower()
    anio_m = re.search(r'(20\d{2})', nombre)
    anio = int(anio_m.group(1)) if anio_m else None
    mes = next((v for k, v in MESES_ES.items() if k in nombre), None)
    return mes, anio


def _norm_col(col) -> str:
    return re.sub(r'\s+', ' ', str(col).strip().upper())


def resolver_ruta_datos(cfg: dict, clave_nuevo: str, clave_legacy: str) -> Path:
    """Retorna la ruta que existe, priorizando nueva estructura."""
    base = Path(cfg['rutas']['base'])
    nuevo  = base / cfg['rutas'][clave_nuevo]
    legacy = base / cfg['rutas'][clave_legacy]

    def _contiene_excels(path: Path) -> bool:
        return path.is_dir() and any(path.glob('*.xlsx'))

    if _contiene_excels(nuevo):
        return nuevo
    if _contiene_excels(legacy):
        return legacy
    if legacy.exists():
        return legacy
    return nuevo  # devuelve nuevo aunque no exista (fallará luego con mensaje claro)


def cargar_pam_produccion(cfg: dict) -> pd.DataFrame:
    """Carga todos los Excel de PAM Producción. Retorna DataFrame diario."""
    ruta = resolver_ruta_datos(cfg, 'pam_produccion_nuevo', 'pam_produccion_legacy')
    mapeo = cfg.get('mapeo_pam_produccion', {})
    archivos = sorted(ruta.glob('*.xlsx'))
    if not archivos:
        log.warning(f"Sin Excel en {ruta}")
        return pd.DataFrame()

    dfs = []
    for f in archivos:
        mes, anio = _mes_anio(f.name)
        if not mes or not anio:
            log.warning(f"No se extrajo mes/año de: {f.name}")
            continue
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            log.error(f"Error {f.name}: {e}")
            continue

        if 'Planta' not in wb.sheetnames:
            log.warning(f"{f.name}: hoja 'Planta' no encontrada")
            wb.close(); continue

        ws = wb['Planta']
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Detectar fila cabecera con SAG
        hdr_idx = col_sag1 = col_sag2 = col_pmc = col_mun = None
        for ri, row in enumerate(rows[:10]):
            rn = [_norm_col(c) for c in row]
            if any('SAG' in c for c in rn):
                hdr_idx = ri
                for ci, c in enumerate(rn):
                    if 'SAG 1' in c: col_sag1 = ci
                    elif 'SAG 2' in c: col_sag2 = ci
                    elif 'MOL 1' in c and '12' in c: col_pmc = ci
                    elif 'MOL 13' in c or 'MUN' in c: col_mun = ci
                break
        if hdr_idx is None:
            log.warning(f"{f.name}: cabecera no encontrada"); continue

        data = []
        for row in rows[hdr_idx + 2:]:
            if not row or row[0] is None: continue
            try: dia = int(row[0])
            except (TypeError, ValueError): continue
            if not 1 <= dia <= 31: continue
            try: fecha = datetime(anio, mes, dia)
            except ValueError: continue

            def sv(idx):
                if idx is None or idx >= len(row): return np.nan
                v = row[idx]
                return float(v) if isinstance(v, (int, float)) else np.nan

            data.append({'fecha': fecha, 'SAG1_prog': sv(col_sag1),
                         'SAG2_prog': sv(col_sag2), 'PMC_prog': sv(col_pmc),
                         'MUN_prog': sv(col_mun)})

        if data:
            dfs.append(pd.DataFrame(data))
            log.info(f"PAM Prod | {f.name}: {len(data)} días")

    if not dfs: return pd.DataFrame()
    df = pd.concat(dfs).sort_values('fecha').drop_duplicates('fecha')
    df['fecha'] = pd.to_datetime(df['fecha'])
    return df.reset_index(drop=True)


def cargar_pam_mantto(cfg: dict) -> pd.DataFrame:
    """Carga PAM Mantto y extrae días con ventana T8. Retorna DataFrame diario."""
    ruta = resolver_ruta_datos(cfg, 'pam_mantto_nuevo', 'pam_mantto_legacy')
    keywords = [k.lower() for k in cfg.get('t8_keywords', ['teniente 8'])]
    archivos = sorted(ruta.glob('*.xlsx'))
    dfs = []

    for f in archivos:
        mes, anio = _mes_anio(f.name)
        if not mes or not anio: continue
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            log.error(f"Error {f.name}: {e}"); continue

        hoja = next((s for s in wb.sheetnames if 'ejecutivo' in s.lower()), None)
        if not hoja:
            wb.close(); continue

        rows = list(wb[hoja].iter_rows(values_only=True))
        wb.close()

        data = []
        for row in rows:
            rs = ' '.join(str(c).lower() for c in row if c is not None)
            if any(kw in rs for kw in keywords):
                for di, val in enumerate(row[6:37]):
                    if isinstance(val, (int, float)) and val > 0:
                        try:
                            data.append({'fecha': datetime(anio, mes, di+1),
                                         'horas_t8': float(val)})
                        except ValueError: pass

        if data:
            dfs.append(pd.DataFrame(data))
            log.info(f"PAM Mantto | {f.name}: {len(data)} días T8")

    if not dfs: return pd.DataFrame(columns=['fecha','horas_t8'])
    df = pd.concat(dfs)
    df['fecha'] = pd.to_datetime(df['fecha'])
    return df.groupby('fecha', as_index=False)['horas_t8'].sum().sort_values('fecha')


def cargar_rendimientos(cfg: dict) -> pd.DataFrame:
    """Carga el archivo de rendimientos reales (5 min). Retorna DataFrame con columnas TPH."""
    base = Path(cfg['rutas']['base'])
    threshold = float(cfg['operacional']['tph_threshold'])

    # Buscar en nuevo directorio primero
    ruta_nuevo = base / cfg['rutas']['rendimientos_nuevo']
    archivos = list(ruta_nuevo.glob('*.xlsx')) if ruta_nuevo.exists() else []

    # Fallback preferente al consolidado historico v2
    if not archivos:
        for candidate in (
            base / 'data' / 'raw' / 'tonelaje_v2.xlsx',
            base / 'data' / 'raw' / 'tonelaje_v2_copy.xlsx',
        ):
            if candidate.exists():
                archivos = [candidate]
                break

    # Fallback al archivo legacy
    if not archivos:
        legacy = base / cfg['rutas']['rendimientos_legacy']
        if legacy.exists():
            archivos = [legacy]
        else:
            log.error(f"No se encontró archivo de rendimientos en {ruta_nuevo} ni en {legacy}")
            return pd.DataFrame()

    dfs = []
    for f in archivos:
        log.info(f"Cargando rendimientos: {f.name}")
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        header = [_norm_col(c) for c in rows[0]]
        col_map = {}
        for i, h in enumerate(header):
            if 'FECHA' in h:                         col_map['fecha'] = i
            elif 'CV_315' in h or 'CV 315' in h:    col_map['CV315'] = i
            elif 'CV_316' in h or 'CV 316' in h:    col_map['CV316'] = i
            elif 'SAG:NIVEL_PILA' in h:             col_map['pila_sag1'] = i
            elif 'SAG2:NIVEL_PILA' in h:            col_map['pila_sag2'] = i
            elif 'REND_TMS_SAG1_PI' in h:           col_map['SAG1'] = i
            elif 'REND_TMS_SAG2_PI' in h:           col_map['SAG2'] = i
            elif 'REND_TMS_PMC' in h:               col_map['PMC'] = i
            elif 'SAG 1' in h or 'SAG1' in h:       col_map['SAG1']  = i
            elif 'SAG 2' in h or 'SAG2' in h:       col_map['SAG2']  = i
            elif 'MUN' in h or 'UNITARIO' in h:     col_map['MUN']   = i
            elif 'CONVENCIONAL' in h or 'PMC' in h: col_map['PMC']   = i

        activos = ['SAG1', 'SAG2', 'MUN', 'PMC']
        data = []
        for row in rows[1:]:
            fecha = row[col_map.get('fecha', 0)]
            if not isinstance(fecha, datetime): continue
            rec = {'fecha': fecha}
            for a in activos:
                idx = col_map.get(a)
                v = row[idx] if idx is not None else None
                rec[a] = float(v) if isinstance(v, (int, float)) else np.nan
            for source_col, target_col in (
                ('CV315', 'CV315'),
                ('CV316', 'CV316'),
                ('pila_sag1', 'pila_sag1'),
                ('pila_sag2', 'pila_sag2'),
            ):
                idx = col_map.get(source_col)
                v = row[idx] if idx is not None else None
                rec[target_col] = float(v) if isinstance(v, (int, float)) else np.nan
            data.append(rec)

        dfs.append(pd.DataFrame(data))

    df = pd.concat(dfs).sort_values('fecha').reset_index(drop=True)
    df['fecha'] = pd.to_datetime(df['fecha'])

    dt_h = cfg['operacional']['dt_min'] / 60.0
    for a in activos:
        df[f'{a}_tph']     = df[a].where(df[a] > threshold, other=0.0)
        df[f'{a}_operando'] = df[a] > threshold
        df[f'{a}_ton']     = df[f'{a}_tph'] * dt_h

    df.drop(columns=activos, inplace=True)
    log.info(f"Rendimientos cargados: {len(df):,} registros")
    return df
