"""
Reorganizacion definitiva del proyecto 07_Rendimientos.
Skill: token_optimization_loop — sin recalcular nada, solo mover y documentar.
"""
import shutil, time, sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(r"c:\Users\jorel038\OneDrive - Codelco\Documentos\AA_CIO_DET\07_Rendimientos")
OUT  = ROOT / "outputs"
t0   = time.time()

# ═══════════════════════════════════════════════════════════════
# 1. CREAR NUEVA ESTRUCTURA DE DIRECTORIOS
# ═══════════════════════════════════════════════════════════════
DIRS = [
    "outputs/reports/01_EDA",
    "outputs/reports/02_EventStudy_T8",
    "outputs/reports/03_Pilas",
    "outputs/reports/04_Autonomia",
    "outputs/reports/05_Modelos",
    "outputs/reports/06_SHAP",
    "outputs/reports/07_Optimizacion_Rates",
    "outputs/reports/08_Modelo_Causal",
    "outputs/reports/09_Modelo_Causal_Final",
    "outputs/reports/99_Historicos",
    "outputs/figures/01_EDA",
    "outputs/figures/02_EventStudy_T8",
    "outputs/figures/03_Gaviota",
    "outputs/figures/04_Pilas",
    "outputs/figures/05_Autonomia",
    "outputs/figures/06_Rates",
    "outputs/figures/07_Modelos",
    "outputs/figures/08_SHAP",
    "outputs/figures/09_Clustering",
    "outputs/figures/10_Drift",
    "outputs/figures/11_Modelo_Causal",
    "outputs/figures/99_Historicos",
    "outputs/excel/datasets",
    "outputs/excel/kpis",
    "outputs/excel/model_registry",
    "outputs/excel/autonomia",
    "outputs/excel/simulaciones",
    "outputs/excel/report_tables",
    "outputs/models/campeones",
    "outputs/models/challengers",
    "outputs/models/descartados",
    "outputs/models/historicos",
    "outputs/dashboards",
    "outputs/presentations",
    "outputs/logs",
    "outputs/archive",
]

print("[1/6] Creando estructura de directorios...")
for d in DIRS:
    (ROOT / d).mkdir(parents=True, exist_ok=True)
print(f"      {len(DIRS)} directorios creados/verificados")


# ═══════════════════════════════════════════════════════════════
# 2. MAPA DE MOVIMIENTOS
# ═══════════════════════════════════════════════════════════════
print("[2/6] Ejecutando movimientos de archivos...")

def mv(src_rel, dst_rel, rename=None):
    src = ROOT / src_rel
    if not src.exists():
        return False
    dst_dir = ROOT / dst_rel
    dst_dir.mkdir(parents=True, exist_ok=True)
    fname = rename if rename else src.name
    dst = dst_dir / fname
    if dst.exists():
        stem, suf = Path(fname).stem, Path(fname).suffix
        dst = dst_dir / (stem + "_dup" + suf)
    shutil.move(str(src), str(dst))
    return True

moved = 0

# ── REPORTS: Markdown ──────────────────────────────────────────
md_moves = [
    ("outputs/reports/Reporte_Ejecutivo.md", "outputs/reports/01_EDA",
     "20260625_EDA_Resumen_Ejecutivo.md"),
    ("outputs/reports/variables_sin_etiqueta.md", "outputs/reports/01_EDA",
     "20260625_EDA_Variables_Sin_Etiqueta.md"),
    ("outputs/reports/resumen_event_study_t8.md", "outputs/reports/02_EventStudy_T8",
     "20260625_EventStudy_T8_Resumen.md"),
    ("outputs/reports/advanced_t8_historical_analysis.md", "outputs/reports/02_EventStudy_T8",
     "20260625_EventStudy_T8_Historico_Avanzado.md"),
    ("outputs/reports/Analisis_T8_Intensidad.md", "outputs/reports/02_EventStudy_T8",
     "20260625_EventStudy_T8_Intensidad.md"),
    ("outputs/reports/reporte_ejecutivo_t8.md", "outputs/reports/02_EventStudy_T8",
     "20260625_EventStudy_T8_Ejecutivo_v1.md"),
    ("outputs/reports/reporte_ejecutivo_t8_maestro.md", "outputs/reports/02_EventStudy_T8",
     "20260625_EventStudy_T8_Ejecutivo_Maestro.md"),
    ("outputs/reports/reporte_tecnico_t8.md", "outputs/reports/02_EventStudy_T8",
     "20260625_EventStudy_T8_Tecnico_v1.md"),
    ("outputs/reports/reporte_tecnico_t8_maestro.md", "outputs/reports/02_EventStudy_T8",
     "20260625_EventStudy_T8_Tecnico_Maestro.md"),
    ("outputs/reports/Narrativa_T8_Claude.md", "outputs/reports/02_EventStudy_T8",
     "20260625_EventStudy_T8_Narrativa.md"),
    ("outputs/reports/monitoreo_efecto_gaviota.md", "outputs/reports/02_EventStudy_T8",
     "20260625_EventStudy_EfectoGaviota_Monitoreo.md"),
    ("outputs/reports/resumen_efecto_gaviota.md", "outputs/reports/02_EventStudy_T8",
     "20260625_EventStudy_EfectoGaviota_Resumen.md"),
    ("outputs/reports/modelo_descarga_pilas_robusto.md", "outputs/reports/03_Pilas",
     "20260625_Pilas_Descarga_Robusto.md"),
    ("outputs/reports/resumen_modelo_dinamico_pilas.md", "outputs/reports/03_Pilas",
     "20260625_Pilas_Modelo_Dinamico.md"),
    ("outputs/reports/autonomia_pilas_report.md", "outputs/reports/04_Autonomia",
     "20260625_Autonomia_Pilas_Report.md"),
    ("outputs/reports/model_improvement_summary.md", "outputs/reports/05_Modelos",
     "20260625_Modelos_Improvement_Summary.md"),
    ("outputs/reports/model_loop_v3_summary.md", "outputs/reports/05_Modelos",
     "20260625_Modelos_Loop_v3_Summary.md"),
    ("outputs/reports/model_master_loop_report.md", "outputs/reports/05_Modelos",
     "20260625_Modelos_Master_Loop.md"),
    ("outputs/reports/modelo_adaptativo_report.md", "outputs/reports/05_Modelos",
     "20260625_Modelos_Adaptativo.md"),
    ("outputs/reports/model_drift_analysis.md", "outputs/reports/05_Modelos",
     "20260625_Modelos_Drift_Analysis.md"),
    ("outputs/reports/model_explainability_report.md", "outputs/reports/06_SHAP",
     "20260625_SHAP_Explainability_v1.md"),
    ("outputs/reports/model_explainability_v3.md", "outputs/reports/06_SHAP",
     "20260625_SHAP_Explainability_v3.md"),
    ("outputs/reports/optimizacion_rates_molienda.md", "outputs/reports/07_Optimizacion_Rates",
     "20260625_OptimizacionRates_Molienda.md"),
    ("outputs/reports/sistema_rt_optimizacion_rates.md", "outputs/reports/07_Optimizacion_Rates",
     "20260625_OptimizacionRates_SistemaRT.md"),
    ("outputs/reports/estrategia_mitigacion_t8.md", "outputs/reports/08_Modelo_Causal",
     "20260625_ModeloCausal_EstrategiaMitigacion.md"),
]
for src, dst, rn in md_moves:
    if mv(src, dst, rn):
        moved += 1

# ── REPORTS: PDFs y PPTX ──────────────────────────────────────
pdf_moves = [
    ("outputs/reports/Informe_Comite_T8.pdf", "outputs/reports/02_EventStudy_T8",
     "20260120_InformeComite_T8_Ejecutivo.pdf"),
    ("outputs/reports/Informe_Comite_v2_T8.pdf", "outputs/reports/02_EventStudy_T8",
     "20260225_InformeComite_T8_Ejecutivo_v2.pdf"),
    ("outputs/reports/Manual_Operacion_Molienda_Basado_En_Datos.pdf",
     "outputs/reports/07_Optimizacion_Rates",
     "20260625_ManualOperacion_Molienda_Ejecutivo.pdf"),
    ("outputs/reports/Anexo_Tecnico_T8.pdf", "outputs/reports/99_Historicos",
     "20260120_AnexoTecnico_T8.pdf"),
    ("outputs/reports/Fase2_Mecanismo_Causal_T8.pdf", "outputs/reports/99_Historicos",
     "20260201_Fase2_MecanismoCausal_T8_Tecnico.pdf"),
    ("outputs/reports/Fase3_Modelo_Pilas_T8.pdf", "outputs/reports/99_Historicos",
     "20260301_Fase3_ModeloPilas_T8_Tecnico.pdf"),
    ("outputs/reports/model_loop_v3_report.pdf", "outputs/reports/99_Historicos",
     "20260501_ModeloLoop_v3_Tecnico.pdf"),
    ("outputs/reports/Informe_Comite_T8.pptx", "outputs/presentations",
     "20260120_InformeComite_T8.pptx"),
]
for src, dst, rn in pdf_moves:
    if mv(src, dst, rn):
        moved += 1

# ── FIGURES: subdirectorios existentes → nueva estructura ─────
fig_dir_moves = [
    ("outputs/figures/event_study",           "outputs/figures/02_EventStudy_T8"),
    ("outputs/figures/advanced_t8_historical","outputs/figures/02_EventStudy_T8"),
    ("outputs/figures/efecto_gaviota",        "outputs/figures/03_Gaviota"),
    ("outputs/figures/pilas",                 "outputs/figures/04_Pilas"),
    ("outputs/figures/modelo_dinamico_pilas", "outputs/figures/04_Pilas"),
    ("outputs/figures/descarga_robusto",      "outputs/figures/04_Pilas"),
    ("outputs/figures/autonomia",             "outputs/figures/05_Autonomia"),
    ("outputs/figures/optimizacion_rates",    "outputs/figures/06_Rates"),
    ("outputs/figures/sistema_rt",            "outputs/figures/06_Rates"),
    ("outputs/figures/model_loop",            "outputs/figures/07_Modelos"),
    ("outputs/figures/model_loop_v3",         "outputs/figures/07_Modelos"),
    ("outputs/figures/model_master",          "outputs/figures/07_Modelos"),
    ("outputs/figures/model_advanced",        "outputs/figures/07_Modelos"),
    ("outputs/figures/modelo_adaptativo",     "outputs/figures/07_Modelos"),
    ("outputs/figures/modelo_hibrido",        "outputs/figures/07_Modelos"),
    ("outputs/figures/decision_operacional",  "outputs/figures/11_Modelo_Causal"),
    ("outputs/figures/prescriptivo",          "outputs/figures/11_Modelo_Causal"),
    ("outputs/figures/ejecutivo",             "outputs/figures/11_Modelo_Causal"),
    ("outputs/figures/fase2",                 "outputs/figures/99_Historicos"),
]
for src_rel, dst_rel in fig_dir_moves:
    src = ROOT / src_rel
    if not src.exists():
        continue
    dst_parent = ROOT / dst_rel
    dst_parent.mkdir(parents=True, exist_ok=True)
    dst = dst_parent / src.name
    if dst.exists():
        dst = dst_parent / (src.name + "_arch")
    shutil.move(str(src), str(dst))
    moved += 1

# ── FIGURES: archivos raiz → subdirectorios semanticos ────────
root_figs = {
    "outputs/figures/01_EDA": [
        "00_F01_Serie_Temporal.png","00_F02_Distribuciones.png","00_F03_Heatmap.png",
        "00_F04_Produccion_Disponibilidad.png","00_F05_Dosis_Respuesta.png",
        "00_F06_PrePost.png","00_F07_STL.png","00_F08_ACF_PACF.png",
        "00_F09_ChangePoints.png","00_F10_Bayesiano.png","00_F11_SHAP_Summary.png",
        "00_F12_SHAP_Dependence.png","00_F13_Indices_Ejecutivos.png",
        "F1_Completitud_Temporal.png","F1_Distribuciones.png","F1_Heatmap_HoraDia.png",
        "F1_Heatmap_Utilizacion.png","F1_Temporal.png","F3_ACF_PACF.png",
        "F3_STL.png","F7_CrossCorr.png","Serie_TPH.png",
    ],
    "outputs/figures/02_EventStudy_T8": [
        "Bayesiano_P_Caida_T8.png","Dosis_Respuesta_T8.png","Efecto_Diferido_T8.png",
        "F2_Dosis_Respuesta.png","Heatmap_Impacto.png","Impacto_T8.png",
        "Ranking_IST8.png","T8_TPH_por_Bucket.png","T8_dosis_respuesta.png",
        "T8_efecto_lags.png",
    ],
    "outputs/figures/04_Pilas": [
        "Balance_Pila_SAG1.png","Balance_Pila_SAG2.png","Consumo_Pilas.png",
        "Correas_Pilas.png","Correas_vs_T8.png","Pilas_TPH.png",
        "Pilas_vs_T8.png","Threshold_SAG1.png","Threshold_SAG2.png",
    ],
    "outputs/figures/07_Modelos": [
        "F6_Bayesiano_Posterior.png","PDP_HorasT8.png","Regresion_TPH_HorasT8.png",
    ],
    "outputs/figures/08_SHAP": [
        "F9_SHAP_Dependence.png","F9_SHAP_Summary.png","SHAP.png",
        "SHAP_Categorias.png","SHAP_Summary_SAG1.png","SHAP_Summary_SAG2.png",
    ],
    "outputs/figures/11_Modelo_Causal": [
        "T8_Cadena_Causal.png",
    ],
}
for dst_rel, files in root_figs.items():
    for f in files:
        if mv(f"outputs/figures/{f}", dst_rel):
            moved += 1

# ── EXCEL ─────────────────────────────────────────────────────
excel_moves = [
    ("outputs/excel/advanced_t8_historical_analysis.xlsx","outputs/excel/datasets"),
    ("outputs/excel/event_study_t8.xlsx",                 "outputs/excel/datasets"),
    ("outputs/excel/eventos_t8_desde_pam.xlsx",           "outputs/excel/datasets"),
    ("outputs/excel/resultados_analisis_t8.xlsx",         "outputs/excel/datasets"),
    ("outputs/excel/resultados_t8.xlsx",                  "outputs/excel/datasets"),
    ("outputs/excel/monitoreo_pre_post_t8.xlsx",          "outputs/excel/datasets"),
    ("outputs/excel/KPI_Molienda.xlsx",                   "outputs/excel/kpis"),
    ("outputs/excel/kpi_autonomia_pilas.xlsx",            "outputs/excel/kpis"),
    ("outputs/excel/IST8_Elasticidad.xlsx",               "outputs/excel/kpis"),
    ("outputs/excel/Analisis_T8.xlsx",                    "outputs/excel/kpis"),
    ("outputs/excel/Analisis_T8_Intensidad.xlsx",         "outputs/excel/kpis"),
    ("outputs/excel/Recuperacion_Post_T8.xlsx",           "outputs/excel/kpis"),
    ("outputs/excel/model_performance_tracking.xlsx",     "outputs/excel/model_registry"),
    ("outputs/excel/model_registry_v2.xlsx",              "outputs/excel/model_registry"),
    ("outputs/excel/model_registry_v3.xlsx",              "outputs/excel/model_registry"),
    ("outputs/excel/modelo_descarga_pilas_robusto.xlsx",  "outputs/excel/autonomia"),
    ("outputs/excel/modelo_dinamico_pilas.xlsx",          "outputs/excel/autonomia"),
    ("outputs/excel/modelo_hibrido_resultados.xlsx",      "outputs/excel/simulaciones"),
    ("outputs/excel/drift_dashboard.xlsx",                "outputs/excel/report_tables"),
    ("outputs/excel/optimizacion_rates_molienda.xlsx",    "outputs/excel/report_tables"),
]
for src, dst in excel_moves:
    if mv(src, dst):
        moved += 1

# ── MODELS ────────────────────────────────────────────────────
champion_models = [
    "outputs/models/capa1_regime_model.pkl",
    "outputs/models/capa2_risk_table.json",
]
for m in champion_models:
    if mv(m, "outputs/models/campeones"):
        moved += 1

# v3 → campeones
v3_dir = ROOT / "outputs/models/v3"
if v3_dir.exists():
    for m in v3_dir.glob("*.pkl"):
        if mv(f"outputs/models/v3/{m.name}", "outputs/models/campeones"):
            moved += 1
    try:
        if not any(v3_dir.iterdir()):
            v3_dir.rmdir()
    except:
        pass

# Challengers: variantes GPU
challenger_models = [
    "outputs/models/catboost_gpu_sag2_tph_mean_v2_gpu.pkl",
    "outputs/models/lightgbm_gpu_sag2_tph_mean_v2_gpu.pkl",
    "outputs/models/xgboost_gpu_sag2_tph_mean_v2_gpu.pkl",
    "outputs/models/histgradientboosting_sag2_tph_mean_v2_gpu.pkl",
    "outputs/models/randomforest_sag2_tph_mean_v2_gpu.pkl",
]
for m in challenger_models:
    if mv(m, "outputs/models/challengers"):
        moved += 1

# Resto de modelos → historicos
for m in list((ROOT / "outputs/models").glob("*.pkl")) + \
         list((ROOT / "outputs/models").glob("*.json")):
    if mv(f"outputs/models/{m.name}", "outputs/models/historicos"):
        moved += 1

print(f"      {moved} archivos/directorios movidos")


# ═══════════════════════════════════════════════════════════════
# 3. MARKDOWN MODELO CAUSAL FINAL
# ═══════════════════════════════════════════════════════════════
print("[3/6] Generando 20260625_Modelo_Causal_Inventario_T8.md...")

CAUSAL_MD = """# Modelo Causal del Inventario de Molienda — Ventanas T8
*Fecha: 2026-06-25 | Division El Teniente — Area Molienda SAG*

---

## 1. Mecanismo causal real del impacto T8

La ventana T8 (Teniente 8) interrumpe el flujo de mineral desde la mina hacia los circuitos de molienda.
El mecanismo es:

```
T8_activo = 1
    -> correa_315 = 0  (sin alimentacion SAG1)
    -> correa_316 = 0  (sin alimentacion SAG2)
    -> dS/dt = Qin - Qout = 0 - Qout  ->  dS/dt < 0
    -> Pila SAG1/SAG2 se drena al ritmo de la molienda
    -> Cuando pila <= umbral_critico -> reduccion forzada de rate
    -> Caida de TPH ("efecto gaviota")
```

La causalidad NO es directa T8->TPH sino MEDIADA por el inventario de pila:

```
T8 -> Pila -> TPH
```

### Hallazgo estructural critico
correa_315 = 0 durante el 49% del tiempo total (no solo en T8).
SAG1 opera sin feed la mitad del tiempo operativo, generando un deficit
cronico de pila que ninguna optimizacion de rate puede resolver sin
intervencion en la disponibilidad de correa_315.

---

## 2. Activo mas vulnerable

SAG1 — Score de vulnerabilidad: 56.4 (vs SAG2: moderado)

| Indicador                  | SAG1          | SAG2          |
|---------------------------|---------------|---------------|
| Autonomia media            | 1.7 h         | 2.6 h         |
| P10 autonomia              | 0.5 h         | 0.2 h         |
| % tiempo auton < 4h        | 100%          | 76.7%         |
| Correa sin feed            | 49% del tiempo| ~30%          |
| Evento mas critico         | 2026-01-02 (-100%) | —       |

---

## 3. Nivel minimo seguro de pila

| Activo | Umbral critico | Umbral operacional |
|--------|---------------|-------------------|
| SAG1   | 15%           | 30%               |
| SAG2   | 18.2%         | 30%               |

---

## 4. Autonomia operacional

```
autonomia_h = (pila_pct - pct_critico) / drain_pct_h
```

| Activo | drain_pct_h | cap_efectiva_ton |
|--------|-------------|-----------------|
| SAG1   | 23.76 %/h   | 4,575 ton       |
| SAG2   | 6.18 %/h    | 32,009 ton      |

Nota: cap_efectiva = TPH_medio / drain_pct_h x 100 (no igual a cap_fisica).

---

## 5. Variables que controlan el rendimiento (SHAP)

| Rank | Variable                     | Activo |
|------|------------------------------|--------|
| 1    | pila_sag1 / pila_sag2        | Ambos  |
| 2    | autonomia_sag1               | SAG1   |
| 3    | correa_315_activa            | SAG1   |
| 4    | horas_sin_correa_315         | SAG1   |
| 5    | t8_activo                    | Ambos  |
| 6    | duracion_h                   | Ambos  |
| 7    | tiempo_a_critico_sag1        | SAG1   |
| 8    | dpila_sag1_dt                | SAG1   |
| 9    | frac_t8_completada           | Ambos  |
| 10   | ratio_pilas                  | Ambos  |

---

## 6. Senales previas a una caida de TPH

1. dpila_dt < -2%/h: velocidad de drenaje se acelera
2. horas_sin_correa_315 > 3h: SAG1 sin reposo prolongado
3. autonomia_sag1 < 2.5h: zona CONSERVADOR -> activar protocolo
4. autonomia_sag1 < 1.0h: zona EMERGENCIA -> reduccion inmediata
5. correa_315 == 0 + t8_activo == 1: doble riesgo simultaneo

Ventana de intervencion util: ~2.3h antes del agotamiento total.

---

## 7. Rate recomendado SAG1 (P90 = 1,454 TPH)

| Regimen     | Condicion                          | Rate (%P90) | TPH          |
|------------|-------------------------------------|------------|--------------|
| EMERGENCIA | auton < 1h o pila < 20%             | 50-64%     | 727-931      |
| CONSERVADOR| T8>=4h o auton < 2.5h              | 58-78%     | 843-1,134    |
| NORMAL     | Operacion estandar                  | 72-95%     | 1,047-1,381  |
| AGRESIVO   | pila > 65% y sin T8                 | 87-105%    | 1,265-1,527  |

---

## 8. Rate recomendado SAG2 (P90 = 2,516 TPH)

Regimen basado en estado SAG2, independiente de autonomia SAG1.

| Regimen     | Condicion                          | Rate (%P90) | TPH          |
|------------|-------------------------------------|------------|--------------|
| EMERGENCIA | auton_SAG2 < 1h o pila_SAG2 < 22%  | 68-82%     | 1,711-2,063  |
| CONSERVADOR| T8>=4h o auton_SAG2 < 2.5h        | 76-94%     | 1,912-2,365  |
| NORMAL     | Operacion estandar                  | 82-100%    | 2,063-2,516  |
| AGRESIVO   | pila_SAG2 > 55% y sin T8           | 90-105%    | 2,264-2,642  |

---

## 9. Cuando reducir carga

Protocolo escalonado:

```
IF autonomia_h < 4h:
    ALERTA: monitorear cada 30 min

IF autonomia_h < 2.5h OR (t8_activo AND duracion_h >= 4):
    CONSERVADOR: reducir rate al 65-80% P90

IF autonomia_h < 1h OR pila < umbral_critico:
    EMERGENCIA: reducir rate al 50-68% P90 + notificar jefatura

IF pila <= umbral_critico AND correa == 0 AND t8_activo:
    DETENCION PREVENTIVA EVALUABLE (ver seccion 10)
```

Regla PRE-VENTANA: iniciar reduccion 24h antes de T8 conocido cuando
pila SAG1 < 65% o pila SAG2 < 55%. Objetivo: llegar a inicio T8 con pila >= 70%.

---

## 10. Cuando evaluar detencion preventiva

| Condicion                                    | Accion recomendada           |
|---------------------------------------------|------------------------------|
| Autonomia proyectada < 2h + T8 >= 8h activo | Evaluar detencion SAG1       |
| pila_SAG1 < 15% + correa_315 = 0 + sin T8  | Pausa para recuperacion pila |
| P(agotamiento 4h) > 80%                     | Detencion preventiva inmediata|
| CV_TPH > 30% por 3+ horas                   | Revision operacional         |

Arbol de decision (regla simplificada):
```
pila_SAG2 <= 22.7%
    baseline_TPH <= 1,986 -> reducir
    baseline_TPH  > 1,986 -> mantener
pila_SAG2 > 22.7%
    pila_SAG1 <= 31.7% -> reducir
    pila_SAG1 en [31.7, 49.4] -> mantener con vigilancia
    pila_SAG1 > 49.4% -> mantener normal
```

---

## Reglas Operacionales (15 reglas)

| #  | Regla                                                              |
|----|-------------------------------------------------------------------|
| 1  | Pre-T8: alcanzar pila >= 70% SAG1 y >= 65% SAG2 antes del inicio |
| 2  | T8 corto (2h): mantener rate > 80% P90 (reserva suficiente)      |
| 3  | T8 largo (>=4h): reducir inmediatamente a CONSERVADOR            |
| 4  | Autonomia < 2.5h: activar CONSERVADOR automaticamente            |
| 5  | Autonomia < 1h: EMERGENCIA + notificar jefatura                  |
| 6  | Post-T8: mantener rate moderado 24h hasta reposicion de pilas    |
| 7  | SAG2 no penalizar por crisis de SAG1 (buffer independiente)      |
| 8  | correa_315 inactiva > 3h: monitoreo reforzado SAG1 c/15 min     |
| 9  | CV_TPH > 25%: investigar causa antes de modificar rate           |
| 10 | pila_SAG1 < 15% + T8 activo: stop SAG1 y esperar feed           |
| 11 | Regimen AGRESIVO solo cuando pila > 65% (SAG1) o > 55% (SAG2)  |
| 12 | No superar 105% P90 en ningun regimen                           |
| 13 | Cambiar rate maximo +-10% P90 por turno (evitar transitorios)   |
| 14 | Registrar cada cambio de regimen en bitacora operacional         |
| 15 | Disparador Power BI: autonomia media turno < 2h -> alerta CIO   |

---

## Balance de Masa (ecuacion diferencial)

```
dS/dt = Qin(t) - Qout(t)

donde:
  S(t)    = inventario pila en ton
  Qin(t)  = flujo correa (ton/h)  — correa_315 o correa_316
  Qout(t) = rate molienda (ton/h) — SAG1_tph o SAG2_tph

Discretizacion (5 min):
  S[t+1] = S[t] + Qin[t] * DT - Qout[t] * DT
  DT = 5/60 h

Agotamiento: S[t] <= S_critico = cap * (pct_critico / 100)
```

---

## Backtesting Sistema RT (abr-jun 2026)

| Activo | Delta TPH | Mejora agotamiento | Delta autonomia |
|--------|-----------|-------------------|----------------|
| SAG1   | -1.4% OK  | -1.1% (limitacion estructural) | +0.01h |
| SAG2   | -0.7% OK  | -3.5% (ver nota)  | +0.24h         |

Capa 1 accuracy: 99.6% | API latencia: ~0.3 s/llamada

Limitacion documentada SAG1: reduccion agotamiento >=20% requiere
mejora en disponibilidad correa_315, no solo optimizacion de rates.

---

## KPIs recomendados para CIO / Power BI

1. Autonomia SAG1/SAG2 (h) — semaforo: >=4h verde, 2-4h amarillo, <2h rojo
2. P10 autonomia (ultimas 24h)
3. % tiempo autonomia < 4h en turno
4. Regimen operacional actual (EMERGENCIA/CONSERVADOR/NORMAL/AGRESIVO)
5. Rate recomendado vs rate operado (delta %)
6. P(agotamiento 4h) en tiempo real
7. Horas sin correa_315
8. Recovery time post-T8 (h al 90% baseline)
9. Eventos de agotamiento en turno (contador)
10. Alerta: autonomia media turno < 2h -> notificacion CIO

---

*Scripts fuente:*
*  src/advanced_t8_historical_analysis.py*
*  src/optimizacion_rates_molienda.py*
*  src/sistema_rt_optimizacion_rates.py*
*Modelos: outputs/models/campeones/capa1_regime_model.pkl*
*Cache: data/cache/advanced_t8_historical_5min.parquet*
*Generado: 2026-06-25*
"""

causal_path = ROOT / "outputs/reports/09_Modelo_Causal_Final/20260625_Modelo_Causal_Inventario_T8.md"
causal_path.write_text(CAUSAL_MD, encoding="utf-8")
print(f"      Generado: {causal_path.name}")


# ═══════════════════════════════════════════════════════════════
# 4. INDICE MAESTRO
# ═══════════════════════════════════════════════════════════════
print("[4/6] Generando Indice_Proyecto_Rendimientos.xlsx...")

HDR_FILL  = PatternFill("solid", fgColor="1F3864")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ROW_FILLS = [PatternFill("solid", fgColor="EBF1F8"), PatternFill("solid", fgColor="FFFFFF")]

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN; cell.border = BORDER

def auto_width(ws, extra=4):
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + extra, 60)

wb_idx = openpyxl.Workbook()
ws = wb_idx.active
ws.title = "Indice"
ws.row_dimensions[1].height = 30
ws.append(["ID","Fecha","Dominio","Tipo","Audiencia","Archivo","Ruta_Relativa"])
style_header(ws)

INDEX = [
    ("R001","2026-06-25","EventStudy_T8","Markdown","Tecnico",
     "20260625_EventStudy_T8_Historico_Avanzado.md","reports/02_EventStudy_T8/"),
    ("R002","2026-06-25","EventStudy_T8","Markdown","Ejecutivo",
     "20260625_EventStudy_T8_Ejecutivo_Maestro.md","reports/02_EventStudy_T8/"),
    ("R003","2026-06-25","EventStudy_T8","Markdown","Tecnico",
     "20260625_EventStudy_T8_Tecnico_Maestro.md","reports/02_EventStudy_T8/"),
    ("R004","2026-06-25","EfectoGaviota","Markdown","Tecnico",
     "20260625_EventStudy_EfectoGaviota_Monitoreo.md","reports/02_EventStudy_T8/"),
    ("R005","2026-01-20","EventStudy_T8","PDF","Ejecutivo",
     "20260120_InformeComite_T8_Ejecutivo.pdf","reports/02_EventStudy_T8/"),
    ("R006","2026-02-25","EventStudy_T8","PDF","Ejecutivo",
     "20260225_InformeComite_T8_Ejecutivo_v2.pdf","reports/02_EventStudy_T8/"),
    ("R007","2026-06-25","Pilas","Markdown","Tecnico",
     "20260625_Pilas_Modelo_Dinamico.md","reports/03_Pilas/"),
    ("R008","2026-06-25","Pilas","Markdown","Tecnico",
     "20260625_Pilas_Descarga_Robusto.md","reports/03_Pilas/"),
    ("R009","2026-06-25","Autonomia","Markdown","Tecnico",
     "20260625_Autonomia_Pilas_Report.md","reports/04_Autonomia/"),
    ("R010","2026-06-25","Modelos","Markdown","Tecnico",
     "20260625_Modelos_Adaptativo.md","reports/05_Modelos/"),
    ("R011","2026-06-25","Modelos","Markdown","Tecnico",
     "20260625_Modelos_Master_Loop.md","reports/05_Modelos/"),
    ("R012","2026-06-25","SHAP","Markdown","Tecnico",
     "20260625_SHAP_Explainability_v3.md","reports/06_SHAP/"),
    ("R013","2026-06-25","Optimizacion_Rates","Markdown","Tecnico",
     "20260625_OptimizacionRates_Molienda.md","reports/07_Optimizacion_Rates/"),
    ("R014","2026-06-25","Optimizacion_Rates","Markdown","Tecnico",
     "20260625_OptimizacionRates_SistemaRT.md","reports/07_Optimizacion_Rates/"),
    ("R015","2026-06-25","Optimizacion_Rates","PDF","Ejecutivo",
     "20260625_ManualOperacion_Molienda_Ejecutivo.pdf","reports/07_Optimizacion_Rates/"),
    ("R016","2026-06-25","Modelo_Causal","Markdown","Ejecutivo",
     "20260625_ModeloCausal_EstrategiaMitigacion.md","reports/08_Modelo_Causal/"),
    ("R017","2026-06-25","ModeloCausal_FINAL","Markdown","Ambos",
     "20260625_Modelo_Causal_Inventario_T8.md","reports/09_Modelo_Causal_Final/"),
    ("H001","2026-01-20","EventStudy_T8","PDF","Tecnico",
     "20260120_AnexoTecnico_T8.pdf","reports/99_Historicos/"),
    ("H002","2026-02-01","EventStudy_T8","PDF","Tecnico",
     "20260201_Fase2_MecanismoCausal_T8_Tecnico.pdf","reports/99_Historicos/"),
    ("H003","2026-03-01","Pilas","PDF","Tecnico",
     "20260301_Fase3_ModeloPilas_T8_Tecnico.pdf","reports/99_Historicos/"),
    ("H004","2026-05-01","Modelos","PDF","Tecnico",
     "20260501_ModeloLoop_v3_Tecnico.pdf","reports/99_Historicos/"),
]
for i, row in enumerate(INDEX):
    ws.append(list(row))
    for cell in ws[ws.max_row]:
        cell.fill = ROW_FILLS[i % 2]; cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

ws.freeze_panes = "A2"
auto_width(ws)
wb_idx.save(ROOT / "outputs/reports/Indice_Proyecto_Rendimientos.xlsx")
print("      Indice_Proyecto_Rendimientos.xlsx generado")


# ═══════════════════════════════════════════════════════════════
# 5. CATALOGO DE FIGURAS
# ═══════════════════════════════════════════════════════════════
print("[5/6] Generando Catalogo_Figuras.xlsx...")

wb_fig = openpyxl.Workbook()
ws2 = wb_fig.active
ws2.title = "Catalogo"
ws2.row_dimensions[1].height = 30
ws2.append(["ID_Fig","Nombre_Archivo","Carpeta","Tema","Script_Fuente","Reporte_Asociado","Descripcion"])
style_header(ws2)

FIG_CATALOG = [
    ("F001","01_EventStudy_SAG1.png","figures/02_EventStudy_T8/event_study","Event Study SAG1",
     "advanced_t8_historical_analysis.py","R002","Impacto T8 en SAG1 — todas duraciones"),
    ("F002","02_EventStudy_SAG2.png","figures/02_EventStudy_T8/event_study","Event Study SAG2",
     "advanced_t8_historical_analysis.py","R002","Impacto T8 en SAG2"),
    ("F003","05_Gaviota_Comparativa.png","figures/02_EventStudy_T8/advanced_t8_historical",
     "Gaviota comparativo","advanced_t8_historical_analysis.py","R004",
     "Forma gaviota SAG1 vs SAG2 vs PMC vs UNITARIO"),
    ("F004","08_Autonomia_Historica_SAG1.png","figures/02_EventStudy_T8/advanced_t8_historical",
     "Autonomia historica SAG1","advanced_t8_historical_analysis.py","R009",
     "Serie historica autonomia SAG1 — media=1.7h"),
    ("F005","09_Autonomia_Historica_SAG2.png","figures/02_EventStudy_T8/advanced_t8_historical",
     "Autonomia historica SAG2","advanced_t8_historical_analysis.py","R009",
     "Serie historica autonomia SAG2 — media=2.6h"),
    ("F006","07_Matriz_Riesgo_SAG1.png","figures/04_Pilas/modelo_dinamico_pilas",
     "Matriz riesgo SAG1","advanced_t8_historical_analysis.py","R008",
     "Riesgo agotamiento SAG1 por nivel inicial x duracion T8"),
    ("F007","08_Matriz_Riesgo_SAG2.png","figures/04_Pilas/modelo_dinamico_pilas",
     "Matriz riesgo SAG2","advanced_t8_historical_analysis.py","R008",
     "Riesgo agotamiento SAG2 por nivel inicial x duracion T8"),
    ("F008","01_Gaviota_SAG1.png","figures/03_Gaviota/efecto_gaviota",
     "Gaviota SAG1","advanced_t8_historical_analysis.py","R004",
     "Perfil normalizado TPH SAG1 por duracion T8"),
    ("F009","02_Gaviota_SAG2.png","figures/03_Gaviota/efecto_gaviota",
     "Gaviota SAG2","advanced_t8_historical_analysis.py","R004",
     "Perfil normalizado TPH SAG2"),
    ("F010","01_Rate_vs_Autonomia_SAG1.png","figures/06_Rates/optimizacion_rates",
     "Rate vs Autonomia SAG1","optimizacion_rates_molienda.py","R013",
     "Curva empirica: rate operado vs autonomia SAG1"),
    ("F011","02_Rate_vs_Autonomia_SAG2.png","figures/06_Rates/optimizacion_rates",
     "Rate vs Autonomia SAG2","optimizacion_rates_molienda.py","R013",
     "Curva empirica: rate operado vs autonomia SAG2"),
    ("F012","08_Heatmap_Rates_Recomendados.png","figures/06_Rates/optimizacion_rates",
     "Heatmap rates","optimizacion_rates_molienda.py","R013",
     "Rate optimo por estado x activo"),
    ("F013","10_Manual_Operacion_Rates.png","figures/06_Rates/optimizacion_rates",
     "Manual operacion rates","optimizacion_rates_molienda.py","R015",
     "Tabla visual rates por regimen — sala de control"),
    ("F014","F01_Regimenes_Operacionales.png","figures/06_Rates/sistema_rt",
     "Regimenes historicos","sistema_rt_optimizacion_rates.py","R014",
     "Distribucion temporal EMERGENCIA/CONSERVADOR/NORMAL/AGRESIVO"),
    ("F015","F02_Modelo_Riesgo_Analitico.png","figures/06_Rates/sistema_rt",
     "Riesgo analitico Capa 2","sistema_rt_optimizacion_rates.py","R014",
     "P(agotamiento) por nivel de pila y rate — MC 200 sim"),
    ("F016","F03_Backtesting_Pilas.png","figures/06_Rates/sistema_rt",
     "Backtesting pilas","sistema_rt_optimizacion_rates.py","R014",
     "pile_sim modelo vs operador abr-jun 2026"),
    ("F017","F04_Backtesting_Rates.png","figures/06_Rates/sistema_rt",
     "Backtesting rates","sistema_rt_optimizacion_rates.py","R014",
     "Rate recomendado vs operado en ventanas T8"),
    ("F018","F05_Metricas_Backtesting.png","figures/06_Rates/sistema_rt",
     "Metricas backtesting","sistema_rt_optimizacion_rates.py","R014",
     "delta_TPH, mejora_agot, delta_auton SAG1 y SAG2"),
    ("F019","SHAP_Summary_Nombres_Operacionales.png","figures/07_Modelos/model_master",
     "SHAP Summary operacional","model_master_loop.py","R012",
     "Importancia SHAP con nomenclatura sala de control"),
    ("F020","SHAP_Dependence_Autonomia.png","figures/07_Modelos/model_master",
     "SHAP Dependence autonomia","model_master_loop.py","R012",
     "Efecto marginal autonomia sobre prediccion TPH"),
    ("F021","T8_Cadena_Causal.png","figures/11_Modelo_Causal",
     "Cadena causal T8","advanced_t8_historical_analysis.py","R017",
     "Diagrama causal: T8 -> Correa -> Pila -> Autonomia -> TPH"),
    ("F022","06_Semaforo_Operacional.png","figures/11_Modelo_Causal/decision_operacional",
     "Semaforo operacional","advanced_t8_historical_analysis.py","R017",
     "Panel de alerta: autonomia, pila, regimen"),
]
for i, row in enumerate(FIG_CATALOG):
    ws2.append(list(row))
    for cell in ws2[ws2.max_row]:
        cell.fill = ROW_FILLS[i % 2]; cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

ws2.freeze_panes = "A2"
auto_width(ws2)
wb_fig.save(ROOT / "outputs/figures/Catalogo_Figuras.xlsx")
print("      Catalogo_Figuras.xlsx generado")


# ═══════════════════════════════════════════════════════════════
# 6. TRAZABILIDAD DE MODELOS
# ═══════════════════════════════════════════════════════════════
print("[6/6] Generando Trazabilidad_Modelos.xlsx...")

wb_tr = openpyxl.Workbook()
ws3 = wb_tr.active
ws3.title = "Trazabilidad"
ws3.row_dimensions[1].height = 30
ws3.append(["ID_Modelo","Nombre_Archivo","Rol","Activo","Dataset",
            "Features_N","Algoritmo","Script","Reporte","Notas"])
style_header(ws3)

TRAZABILIDAD = [
    ("M001","capa1_regime_model.pkl","CAMPEON","SAG1+SAG2",
     "advanced_t8_historical_5min.parquet","19","LightGBM",
     "sistema_rt_optimizacion_rates.py","R014",
     "Clasificador regimen 4 clases. acc=99.6%. Train hasta 2026-04-01"),
    ("M002","capa2_risk_table.json","CAMPEON","SAG1+SAG2",
     "advanced_t8_historical_5min.parquet","—","Monte Carlo analitico",
     "sistema_rt_optimizacion_rates.py","R014",
     "P(agotamiento) lookup table 200 sim. Sin ML. Capa 2 del sistema RT"),
    ("M003","Ridge_core.pkl","CAMPEON","SAG2",
     "advanced_t8_event_windows.parquet","12","Ridge",
     "model_master_loop.py","H004",
     "Modelo base core — regularizacion L2. Campeon accuracy"),
    ("M004","Ridge_autonomia.pkl","CAMPEON","SAG1+SAG2",
     "advanced_t8_event_windows.parquet","8","Ridge",
     "model_master_loop.py","H004",
     "Target: autonomia_h. Usado en reglas operacionales"),
    ("M005","ElasticNet_core.pkl","CAMPEON","SAG2",
     "advanced_t8_event_windows.parquet","12","ElasticNet",
     "model_master_loop.py","H004",
     "Modelo alternativo L1+L2. Challenger al Ridge_core"),
    ("M006","lightgbm_sag2_tph_mean_v01.pkl","HISTORICO","SAG2",
     "rendimientos_clean.parquet","8","LightGBM",
     "model_loop.py","—",
     "Primera version LightGBM — superado por versiones posteriores"),
    ("M007","gradientboosting_sag2_tph_mean_v17.pkl","HISTORICO","SAG2",
     "advanced_t8_event_windows.parquet","15","GradientBoosting",
     "model_loop_v3.py","H004",
     "Ultima version GBM. Superado por LightGBM GPU en velocidad"),
    ("M008","catboost_sag2_tph_mean_v13.pkl","HISTORICO","SAG2",
     "advanced_t8_event_windows.parquet","15","CatBoost",
     "model_loop_v3.py","H004",
     "CatBoost con early stopping. R2=0.71 en test"),
    ("M009","lightgbm_gpu_sag2_tph_mean_v2_gpu.pkl","CHALLENGER","SAG2",
     "advanced_t8_event_windows.parquet","15","LightGBM-GPU",
     "model_advanced.py","—",
     "Version GPU. Mismo R2 que CPU. Reserva para escalar"),
    ("M010","xgboost_gpu_sag2_tph_mean_v2_gpu.pkl","CHALLENGER","SAG2",
     "advanced_t8_event_windows.parquet","15","XGBoost-GPU",
     "model_advanced.py","—",
     "Challenger XGBoost — validar drift antes de promover"),
]
for i, row in enumerate(TRAZABILIDAD):
    ws3.append(list(row))
    fill_color = (
        "D4EFDF" if row[2] == "CAMPEON" else
        "FEF9E7" if row[2] == "CHALLENGER" else "F2F3F4"
    )
    for cell in ws3[ws3.max_row]:
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

ws3.freeze_panes = "A2"
auto_width(ws3)
wb_tr.save(ROOT / "outputs/reports/Trazabilidad_Modelos.xlsx")
print("      Trazabilidad_Modelos.xlsx generado")


# ═══════════════════════════════════════════════════════════════
# RESUMEN FINAL
# ═══════════════════════════════════════════════════════════════
elapsed = time.time() - t0
total_files = sum(1 for _ in OUT.rglob("*") if _.is_file())
total_dirs  = sum(1 for _ in OUT.rglob("*") if _.is_dir())

print()
print("=" * 65)
print("  REORGANIZACION COMPLETADA")
print("=" * 65)
print(f"  Archivos movidos/directorios reorganizados : {moved}")
print(f"  Total archivos en outputs ahora            : {total_files}")
print(f"  Total carpetas en outputs ahora            : {total_dirs}")
print(f"  Tiempo ejecucion                           : {elapsed:.1f}s")
print()
print("  Entregables generados:")
print("    reports/09_Modelo_Causal_Final/20260625_Modelo_Causal_Inventario_T8.md")
print("    reports/Indice_Proyecto_Rendimientos.xlsx")
print("    figures/Catalogo_Figuras.xlsx")
print("    reports/Trazabilidad_Modelos.xlsx")
print()
print("  Modelos campeon:")
print("    models/campeones/capa1_regime_model.pkl  (LightGBM, acc=99.6%)")
print("    models/campeones/capa2_risk_table.json   (Monte Carlo analitico)")
print("    models/campeones/Ridge_core.pkl")
print("    models/campeones/ElasticNet_core.pkl")
print("=" * 65)
